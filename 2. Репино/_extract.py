"""
Сбор данных по жизненному циклу проекта МСГ_RBI Репино Санаторий.

Жизненный цикл (этапы) лежит в листе «МСГ, ГПР», колонка 38 (ПЛАН/ФАКТ):
  РД (рабочая документация)         — План РД / Факт РД
  Пакет (комплект документов)       — План П / Факт П
  Тендер                            — План Т / Факт Т
  Договор                           — План Д / Факт Д
  Финансирование (аванс)            — План Ф / Факт Ф
  Мобилизация (подрядчик + МТР)     — План М / Факт М
  СМР (фактическое выполнение)      — План / Факт

Колонка 33 — % завершения, 24 — наименование, 25 — здание, 40/41 — даты.
Раздел задаётся отдельной строкой-заголовком (имя в к.24, в к.38 None).

Лист «РС» — расценочный лист по 31 лоту (стоимости).
"""
import openpyxl
from collections import defaultdict
from pathlib import Path
import json

XLSX = Path(r"C:\Авраменко\Claude Code Projects\МСГ\2. Репино\МСГ_RBI Репино Санаторий.xlsx")
OUT = Path(r"C:\Авраменко\Claude Code Projects\МСГ\2. Репино\_lifecycle.json")

LIFECYCLE = [
    ("РД (Рабочая документация)",        "План РД", "Факт РД", "Ниязов Р."),
    ("Пакеты документов",                "План П",  "Факт П",  "Кулешова Н."),
    ("Тендеры",                          "План Т",  "Факт Т",  "Лисица Д."),
    ("Договоры",                         "План Д",  "Факт Д",  "Островская Т."),
    ("Финансирование (аванс)",           "План Ф",  "Факт Ф",  "Гончарова А."),
    ("Мобилизация подрядчиков и МТР",    "План М",  "Факт М",  "Леонтьев А."),
    ("СМР (выполнение работ)",           "План",    "Факт",    "Горошков / Мещеряков / Лемента"),
]

SUBHEADER_PREFIXES = (
    "Выдача РД", "Пакет -", "Тендер -", "Заключение договора",
    "Финансирование", "Мобилизация подрядчика", "Мобилизация на объекте",
)

def to_pct(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v) * 100.0
    s = str(v).strip().replace(",", ".").replace("%", "")
    try: return float(s)
    except ValueError: return None

def classify(p):
    if p is None: return "no_data"
    if p >= 99.5: return "done"
    if p <= 0.5:  return "not_started"
    return "in_progress"

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["МСГ, ГПР"]

    # buckets[stage_label] = list of {section, name, building, pct, start, end}
    buckets = defaultdict(list)
    cur_section = None

    for row in ws.iter_rows(min_row=5, values_only=True):
        if len(row) < 42: continue
        name = row[23]
        bld  = row[24]
        pf   = row[37]
        pct  = to_pct(row[32])
        d_st = row[39]
        d_en = row[40]

        if name and pf is None:
            t = str(name).strip()
            if not t.startswith(SUBHEADER_PREFIXES):
                cur_section = t
            continue

        if pf is None:
            continue

        pf_s = str(pf).strip()
        rec = {
            "section": cur_section,
            "name": str(name).strip() if name else None,
            "building": str(bld).strip() if bld else None,
            "pct": pct,
            "start": str(d_st)[:10] if d_st else None,
            "end":   str(d_en)[:10] if d_en else None,
        }
        buckets[pf_s].append(rec)

    # Сводка по жизненному циклу — берём «факт-этапы»
    lifecycle = []
    for label, plan_key, fact_key, resp in LIFECYCLE:
        items = buckets.get(fact_key, [])
        total = len(items)
        cnt = {"done": 0, "in_progress": 0, "not_started": 0, "no_data": 0}
        for it in items:
            cnt[classify(it["pct"])] += 1
        lifecycle.append({
            "stage": label,
            "responsible": resp,
            "total": total,
            **cnt,
            "pct_done": (100.0 * cnt["done"] / total) if total else 0.0,
        })

    # Разрез СМР по разделам
    sections_smr = defaultdict(lambda: {"total": 0, "done": 0, "in_progress": 0, "not_started": 0, "no_data": 0})
    for it in buckets.get("Факт", []):
        sec = it["section"] or "(без раздела)"
        s = sections_smr[sec]
        s["total"] += 1
        s[classify(it["pct"])] += 1

    # Разрез СМР по корпусам
    by_building = defaultdict(lambda: {"total": 0, "done": 0, "in_progress": 0, "not_started": 0, "no_data": 0})
    for it in buckets.get("Факт", []):
        b = it["building"] or "(общее)"
        s = by_building[b]
        s["total"] += 1
        s[classify(it["pct"])] += 1

    # Активные СМР работы (0 < % < 100)
    active_smr = []
    for plan, fact in zip(buckets.get("План", []), buckets.get("Факт", [])):
        p = fact["pct"]
        if p is not None and 0.5 < p < 99.5:
            active_smr.append({
                "section": fact["section"],
                "name": plan["name"] or fact["name"],
                "building": plan["building"] or fact["building"],
                "plan_start": plan["start"], "plan_end": plan["end"],
                "fact_pct": p,
            })
    active_smr.sort(key=lambda x: (x["section"] or "", x["building"] or "", x["name"] or ""))

    # Лист «РС» — стоимости по лотам
    ws2 = wb["РС"]
    lots = defaultdict(lambda: {"positions": 0, "cost_total": 0.0})
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 22: continue
        lot = row[0]
        cost = row[21]
        if not lot: continue
        lots[lot]["positions"] += 1
        if isinstance(cost, (int, float)):
            lots[lot]["cost_total"] += float(cost)

    lot_summary = sorted(
        ({"lot": k, **v} for k, v in lots.items()),
        key=lambda x: x["cost_total"], reverse=True
    )
    total_cost = sum(l["cost_total"] for l in lot_summary)
    total_positions = sum(l["positions"] for l in lot_summary)

    payload = {
        "lifecycle": lifecycle,
        "sections_smr": dict(sections_smr),
        "by_building": dict(by_building),
        "active_smr": active_smr,
        "lots": lot_summary,
        "total_cost": total_cost,
        "total_positions": total_positions,
        "raw_counts": {k: len(v) for k, v in buckets.items()},
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Печать сводки
    print("=" * 80)
    print(f"{'ЭТАП':<35} {'Всего':>6} {'Заверш':>7} {'В раб':>6} {'Не нач':>7} {'%':>5}")
    print("-" * 80)
    for s in lifecycle:
        print(f"{s['stage']:<35} {s['total']:>6} {s['done']:>7} {s['in_progress']:>6} {s['not_started']:>7} {s['pct_done']:>5.1f}")
    print()
    print(f"Активных СМР: {len(active_smr)}")
    print(f"Лотов в смете: {len(lot_summary)}, позиций: {total_positions}, общая стоимость: {total_cost:,.0f} руб")
    print(f"\nJSON сохранён: {OUT}")

if __name__ == "__main__":
    main()
