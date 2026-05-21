import openpyxl
from collections import Counter
from pathlib import Path

p = Path(r"C:\Авраменко\Claude Code Projects\МСГ\2. Репино\МСГ_RBI Репино Санаторий.xlsx")
wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
ws = wb["РС"]

print(f"РС: {ws.max_row} rows × {ws.max_column} cols")

# header
hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print("Header:")
for ci, v in enumerate(hdr, 1):
    print(f"  c{ci}: {v!r}")

# Уникальные значения в первой колонке (наименование лота)
lots = Counter()
contractors = Counter()
sections = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]: lots[row[0]] += 1
    if len(row) > 6 and row[6]: sections[row[6]] += 1
    # подрядчик может быть в каких-то колонках; найду эвристически

print(f"\nЛотов уникальных: {len(lots)}")
for k, v in lots.most_common(30):
    print(f"  {v:>4}  {k}")

print(f"\nРазделов в РС (col7): {len(sections)}")
for k, v in sections.most_common(30):
    print(f"  {v:>4}  {k}")
