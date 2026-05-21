"""
fill_1_0_from_gsheet.py — заполняет «Таблица 1-0 Репино.xlsx» значениями 1/0
на основе листа «МСГ, ГПР» из выгрузки Google Sheets Репино.

Два источника готовности (объединяются по OR):
1) Парные строки «Зак. ГПР» / «Факт ГПР» — укрупнённая готовность:
   - «ФУНДАМЕНТНАЯ ПЛИТА»                  → Цоколь.Плита
   - «МОНОЛИТНЫЕ Ж/Б СТЕНЫ НИЖЕ ОТМ. 0.000» → Цоколь.Стены
   - «МОНОЛИТНЫЕ Ж/Б СТЕНЫ ВЫШЕ ОТМ. 0.000 N этаж» → (Корпус, "N этаж"|"Мансарда", Стены)
2) Парные строки «План (Бетонирование …)» / «Факт» — детальная:
   - по конкретной плите (ПП0/ПП1/ПП2/ПП3, плита покрытия +13.400/+16.700, плиты рамп, лестницы)

Готовность ячейки:
- 1, если AO (план/факт окончание) парной Факт-строки заполнено (не пусто, не "-")
- 0, если в плане есть строка для этой ячейки, но Факт.AO пуст
- пусто, если в плане строки нет (конструктив на этой ячейке не предусмотрен)

Использование:
    python fill_1_0_from_gsheet.py --gsheet <downloaded.xlsx>
    python fill_1_0_from_gsheet.py            # автоматически скачает gsheet
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CONFIG_PATH = ROOT / "config.json"

# Координаты колонок в листе «МСГ, ГПР» (1-based)
COL_D = 4      # Раздел
COL_E = 5      # Подраздел
COL_X = 24     # Наименование
COL_Y = 25     # Здание (К1..К6)
COL_AL = 38    # План/Факт
COL_AN = 40    # план/факт начало
COL_AO = 41    # план/факт окончание

CORPUSES = ("К1", "К2", "К3", "К4", "К5", "К6")


def is_done(val) -> bool:
    """Считается ли значение AO заполненным как факт-завершение."""
    if val is None:
        return False
    if isinstance(val, str):
        v = val.strip()
        if not v or v == "-":
            return False
        return True
    return True  # datetime, число — заполнено


def resolve_building(cfg: dict, key: str) -> dict:
    b = cfg["buildings"][key]
    if "extends" in b:
        tpl = cfg["templates"][b["extends"]]
        merged = dict(tpl)
        for k, v in b.items():
            if k != "extends":
                merged[k] = v
        return merged
    return b


# ── маппинг строки X на (этаж, колонка) ───────────────────────────────────
RE_PP_N = re.compile(r"\bпп\s*(\d+)\b", re.IGNORECASE)
RE_ELEV = re.compile(r"([+\-−]?\s*\d+[.,]\d{2,3})")
RE_STEN_N_ETAJ = re.compile(r"стен[аы]?\s*(\d+)[\s\-]*(?:[оеы]?го|ого)?\s*этажа", re.IGNORECASE)
RE_FUND = re.compile(r"фундамент", re.IGNORECASE)
RE_RAMPA = re.compile(r"плит[аы]\s*рампы\s*(ПР\d-?\d?)", re.IGNORECASE)
RE_LESTN_LK = re.compile(r"лестниц(?:ы|ей)?\s+(ЛК\d|ЛМ\d-?\d?)", re.IGNORECASE)
RE_LESTN_PL = re.compile(r"лестничн\w+\s+площадок", re.IGNORECASE)
RE_KROVL = re.compile(r"\bкровл", re.IGNORECASE)


def parse_elev_floor(x: str, b_resolved: dict, b_key: str) -> Optional[str]:
    """По наименованию и отметке вернуть имя этажа из config (или None)."""
    floors = [f["name"] for f in b_resolved["floors"]]
    # Сначала по PP-номеру
    m = RE_PP_N.search(x)
    if m:
        pp_n = int(m.group(1))
        if pp_n == 0:
            return "Цоколь"
        # отметка
        elev = None
        em = RE_ELEV.search(x.split("ПП")[-1] if "ПП" in x.upper() else x)
        if em:
            try:
                elev = float(em.group(1).replace(",", ".").replace("−", "-").replace(" ", ""))
            except ValueError:
                pass
        # типовое сопоставление: ПП1 над 1-м этажом → "1 этаж.Перекрытие"
        # Исключение: «ПП2» в gsheet К3-К6 встречается ДВАЖДЫ — +6.500 и +9.800.
        # Та, что +9.800, физически = ПП3 (плита над 3 этажом).
        if pp_n == 2 and elev is not None and abs(elev - 9.8) < 0.1:
            return "3 этаж" if "3 этаж" in floors else None
        if pp_n == 2 and elev is not None and abs(elev - 6.5) < 0.5:
            return "2 этаж" if "2 этаж" in floors else None
        target = f"{pp_n} этаж"
        return target if target in floors else None
    return None


def map_x_to_cell(x: str, d: str, e: str, b_resolved: dict, b_key: str) -> Optional[tuple[str, str]]:
    """
    Вернуть (имя_этажа, ключ_колонки) или None если строка не маппится.
    Ключи колонок: plate, wall, ceiling, stair, ramp.
    """
    x_low = (x or "").lower()
    d_low = (d or "").lower() if d else ""
    e_low = (e or "").lower() if e else ""

    # Фундамент
    if "фундамент" in x_low:
        return ("Цоколь", "plate")

    # Стены подвала: D = «Монолит ниже 0», E = «Монолитные стены»
    if "монолит ниже 0" in d_low and "монолитные стены" in e_low:
        return ("Цоколь", "wall")

    # Перекрытия (D или E содержит «перекрытия» либо строка содержит ПП)
    is_perekr = ("перекрытие" in x_low) or ("перекрытия" in (d_low + " " + e_low)) or RE_PP_N.search(x or "")

    # Плита покрытия с отметкой
    if "плит" in x_low and "покрыти" in x_low:
        em = RE_ELEV.search(x)
        if em:
            try:
                elev = float(em.group(1).replace(",", ".").replace("−", "-").replace(" ", ""))
            except ValueError:
                elev = None
            if elev is not None:
                if abs(elev - 13.4) < 0.2:
                    # для К3-К6 это перекрытие 4-го этажа
                    return ("4 этаж", "ceiling") if "4 этаж" in [f["name"] for f in b_resolved["floors"]] else None
                if abs(elev - 16.7) < 0.2:
                    # плита покрытия шахт — на «Мансарду» в колонку Плита
                    return ("Мансарда", "plate") if "Мансарда" in [f["name"] for f in b_resolved["floors"]] else None
        # без отметки: для К1 — это ПП2 (на «2 этаж»), для К2 — ПП3 (на «3 этаж»)
        if b_key == "К1":
            return ("2 этаж", "ceiling")
        if b_key == "К2":
            return ("3 этаж", "ceiling")

    # Перекрытие по номеру ПП<N> и отметке (для К3-К6 коррекция ПП2/+9.800 → 3 этаж)
    if is_perekr and RE_PP_N.search(x or ""):
        f = parse_elev_floor(x, b_resolved, b_key)
        if f:
            return (f, "ceiling")

    # Стены этажа N
    m = RE_STEN_N_ETAJ.search(x_low) if x else None
    if m:
        n = int(m.group(1))
        floors = [f["name"] for f in b_resolved["floors"]]
        # для К3-К6 «5-ый этаж» → «Мансарда»
        if n == 5 and "Мансарда" in floors:
            return ("Мансарда", "wall")
        target = f"{n} этаж"
        return (target, "wall") if target in floors else None

    # Лестница (ЛК / ЛМ)
    m = RE_LESTN_LK.search(x or "")
    if m:
        name = m.group(1)
        # ищем в b_resolved.stairs строку, содержащую этот код
        for s in b_resolved.get("stairs", []):
            if name.upper() in s.upper():
                return (f"stair:{s}", "stair")
        # если не нашли, попадаем в первую «Монолитные лестницы подвала» или None
        return None

    # Лестничные площадки → «Монолитные лестницы подвала»
    if RE_LESTN_PL.search(x or ""):
        for s in b_resolved.get("stairs", []):
            if "подвал" in s.lower():
                return (f"stair:{s}", "stair")
        return None

    # Рампа (плита рампы ПР…)
    m = RE_RAMPA.search(x or "")
    if m:
        code = m.group(1).upper()
        # `ПР1` в gsheet, в config «ПР0-1», «ПР1-1» — мапим по числу
        m_num = re.search(r"ПР(\d+)", code)
        if m_num:
            ramp_n = m_num.group(1)
            for r in b_resolved.get("ramps", []):
                if f"ПР{ramp_n}-" in r.upper() or f"ПР{ramp_n} " in r:
                    return (f"ramp:{r}", "ramp")
                # fallback: первое слово начинается с этого префикса
                if r.startswith(f"ПР{ramp_n}"):
                    return (f"ramp:{r}", "ramp")
        return None

    # Кровля
    if RE_KROVL.search(x or ""):
        return ("Кровля", "ceiling")

    return None


def collect_done(gsheet_path: Path, cfg: dict) -> dict[tuple[str, str, str], int]:
    """
    Вернуть карту (corpus, floor_or_subrowkey, col_key) → 1 или 0.
    floor_or_subrowkey — обычное имя этажа, либо "stair:<name>", "ramp:<name>".
    """
    wb = load_workbook(gsheet_path, data_only=True)
    ws = wb["МСГ, ГПР"]
    result: dict[tuple[str, str, str], int] = {}

    # Резолвим корпуса один раз
    resolved = {k: resolve_building(cfg, k) for k in cfg["buildings"].keys()}

    # Источник 1: парные строки
    #   План «Бетонирование …»  (AL = «План»)
    #   Факт                     (AL = «Факт»)  — следующая строка
    # Для всего, кроме кровли — это единственный источник.
    for r in range(2, ws.max_row + 1):
        x = ws.cell(r, COL_X).value
        if not x or not str(x).lower().startswith("бетонирование"):
            continue
        y = ws.cell(r, COL_Y).value
        if y not in CORPUSES:
            continue
        al = ws.cell(r, COL_AL).value
        if al != "План":
            continue
        # парная Факт-строка
        fact_al = ws.cell(r + 1, COL_AL).value
        if fact_al != "Факт":
            continue
        fact_ao = ws.cell(r + 1, COL_AO).value
        d = ws.cell(r, COL_D).value
        e = ws.cell(r, COL_E).value
        cell = map_x_to_cell(x, d, e, resolved[y], y)
        if cell is None:
            continue
        floor_name, col_key = cell
        key = (y, floor_name, col_key)
        val = 1 if is_done(fact_ao) else 0
        # Если уже было записано «1» для этой ячейки от другой подработы — оставляем 1.
        # Если же было «0», а пришла «1» — обновляем на 1 (любая выполненная подработа = ячейка готова).
        if key in result:
            if val == 1:
                result[key] = 1
        else:
            result[key] = val

    # Источник 2: «Кровля» — раздел не из бетонирования, берём по ГПР.
    #   AL = «Зак. ГПР», X содержит «кровли» (металлическая К1/К2 или эксплуатируемая К3-К6)
    #   парный AL = «Факт ГПР» — обычно r+2
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        y = ws.cell(r, COL_Y).value
        if y not in CORPUSES:
            continue
        x = ws.cell(r, COL_X).value or ""
        x_low = x.lower()
        if "кровл" not in x_low or "устройств" not in x_low:
            continue
        # ищем Факт ГПР в r+1..r+4
        fact_ao = None
        for off in (1, 2, 3, 4):
            if ws.cell(r + off, COL_AL).value == "Факт ГПР":
                fact_ao = ws.cell(r + off, COL_AO).value
                break
        key = (y, "Кровля", "ceiling")
        val = 1 if is_done(fact_ao) else 0
        if key in result:
            if val == 1:
                result[key] = 1
        else:
            result[key] = val

    return result


# ── запись в шаблон ─────────────────────────────────────────────────────────
def find_cell_in_template(ws_t, b_key: str, floor_name: str, col_key: str,
                          col_letters: dict[str, str]) -> Optional[str]:
    """
    Пробежать по строкам шаблона, найти секцию корпуса и строку нужного этажа/субблока.
    Возвращает координату ячейки (например, 'C7') или None.
    """
    in_section = False
    in_stairs = False
    in_ramps = False

    target_kind = "floor"
    target_name = floor_name
    if floor_name.startswith("stair:"):
        target_kind = "stair"
        target_name = floor_name.split(":", 1)[1]
    elif floor_name.startswith("ramp:"):
        target_kind = "ramp"
        target_name = floor_name.split(":", 1)[1]

    for r in range(1, ws_t.max_row + 1):
        a = ws_t.cell(r, 1).value
        if not a:
            continue
        a_str = str(a).strip()
        # начало секции корпуса
        if a_str == f"Корпус {b_key}":
            in_section = True
            in_stairs = False
            in_ramps = False
            continue
        if in_section:
            # уход в следующую секцию
            if a_str.startswith("Корпус ") and a_str != f"Корпус {b_key}":
                return None
            if a_str.startswith("Итог по корпусу"):
                return None
            # подзаголовки
            if a_str == "Лестницы":
                in_stairs = True
                in_ramps = False
                continue
            if a_str == "Рампы":
                in_stairs = False
                in_ramps = True
                continue
            # совпадение по имени. В шаблоне может стоять «Цоколь  (отм. …)» — берём префикс до « (отм.».
            stripped = a_str.strip()
            stripped_name = stripped.split("(отм.")[0].strip() if "(отм." in stripped else stripped
            if target_kind == "floor" and not in_stairs and not in_ramps:
                if stripped_name == target_name:
                    return f"{col_letters[col_key]}{r}"
            elif target_kind == "stair" and in_stairs:
                if stripped == target_name:
                    return f"{col_letters[col_key]}{r}"
            elif target_kind == "ramp" and in_ramps:
                if stripped == target_name:
                    return f"{col_letters[col_key]}{r}"
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", type=Path, default=CONFIG_PATH)
    ap.add_argument("--gsheet", type=Path, default=None,
                    help="Путь к скачанному xlsx с листом «МСГ, ГПР». Если не указан — скачаем.")
    ap.add_argument("--template", type=Path, default=None,
                    help="Путь к шаблону «Таблица 1-0 Репино.xlsx» (по умолчанию — из config.output).")
    args = ap.parse_args()

    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    with args.config.open(encoding="utf-8") as f:
        cfg = json.load(f)

    if args.template:
        tpl_path = args.template
    else:
        tpl_path = Path(cfg["output"]["dir"]) / cfg["output"]["filename"]

    # 1) Скачать gsheet при необходимости
    if args.gsheet is None:
        import subprocess
        gid = cfg["source"]["gdrive_file_id"]
        tmp = Path(tempfile.gettempdir()) / f"repino_msg_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        downloader = Path(r"C:\Авраменко\Claude Code Projects\МСГ\scripts\gsheet_download_xlsx.py")
        print(f"[1] Скачиваю gsheet -> {tmp}")
        subprocess.check_call(["python", str(downloader), "--file-id", gid, "--out", str(tmp)])
        gsheet_path = tmp
    else:
        gsheet_path = args.gsheet
        print(f"[1] Используем переданный gsheet: {gsheet_path}")

    # 2) Собрать карту готовности
    print(f"[2] Читаю «МСГ, ГПР» и собираю карту готовности...")
    done_map = collect_done(gsheet_path, cfg)
    print(f"    Найдено целевых ячеек: {len(done_map)}")
    n_ones = sum(1 for v in done_map.values() if v == 1)
    print(f"    Готово (=1): {n_ones}, не готово (=0): {len(done_map) - n_ones}")

    # 3) Открыть шаблон, найти ячейки, проставить значения
    print(f"[3] Записываю значения в шаблон: {tpl_path}")
    wb = load_workbook(tpl_path)
    ws = wb["Таблица 1-0"]

    # Карта col_key → буква колонки шаблона.
    # Шаблон: A=имя, B=Плита, C=Стены, D=Перекрытие, E=Лестницы, F=Рампа, G=Итог
    col_letters = {"plate": "B", "wall": "C", "ceiling": "D", "stair": "E", "ramp": "F"}

    n_written = 0
    n_missing = 0
    skipped = []
    for (corpus, floor, ck), val in sorted(done_map.items()):
        coord = find_cell_in_template(ws, corpus, floor, ck, col_letters)
        if coord is None:
            n_missing += 1
            skipped.append((corpus, floor, ck, val))
            continue
        ws[coord].value = val
        n_written += 1

    print(f"    Записано: {n_written}, не сопоставлено: {n_missing}")
    if skipped:
        print("    Не сопоставленные (первые 20):")
        for s in skipped[:20]:
            print(f"      {s}")

    # 4) Сохранить
    wb.save(tpl_path)
    print(f"[OK] Сохранено: {tpl_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
