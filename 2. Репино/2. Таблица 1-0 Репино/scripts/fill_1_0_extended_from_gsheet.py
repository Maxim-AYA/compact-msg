"""
fill_1_0_extended_from_gsheet.py — заполняет «Таблица 1-0 расширенная Репино.xlsx»
данными из листа «МСГ, ГПР» Репино.

Шаг 1 (текущий): монолит + кровля — то же, что уже работает в существующей `fill_1_0_from_gsheet.py`.
Следующие шаги (TODO):
  - Зона А «До 0»: Подгот./Котлован/Шпунт/БК/ГИ гор./ГИ верт./Обр.засыпка/Утепл. ниже 0
  - Зона В «По корпусу»: Окна Al/Витражи/Фасад/Лифты/МК/Двери/Сборн.каркас
  - Зона Б2 «Отделка/инж.»: парсинг X-строк «План <что-то> N этажа»

Принципы:
- Не трогаем формулы ИТОГО (записываем только в data-ячейки).
- Не пишем в серые «—» N/A-ячейки.
- Источники готовности — то же, что в базовой версии:
  парный «Факт» под «План Бетонирование …» (для монолита);
  парный «Факт ГПР» под «Зак. ГПР Устройство (металлической|эксплуатируемой) кровли» (для строки «Кровля»).

Использование:
    python fill_1_0_extended_from_gsheet.py --gsheet <downloaded.xlsx>
    python fill_1_0_extended_from_gsheet.py            # автоматически скачает gsheet
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CONFIG_PATH = ROOT / "config.json"
DEFAULT_TEMPLATE_NAME = "Таблица 1-0 расширенная Репино.xlsx"

# Колонки исходного gsheet
COL_D = 4
COL_E = 5
COL_X = 24
COL_Y = 25
COL_AL = 38
COL_AN = 40
COL_AO = 41

CORPUSES = ("К1", "К2", "К3", "К4", "К5", "К6")

# Колонки расширенного шаблона (1-based)
# A=Корпус, B=Этаж/Отметка
# Зона A «До 0»: C..J  (8 колонок)
# Зона Б1 «Конструктив»: K..O (5 колонок)
# Зона В «По корпусу»: P..W (8 колонок)
# Зона Б2 «Отделка/инж.»: X..AF (9 колонок)
# Итог: AG
TPL_COLS = {
    # Зона A «До 0»
    "prep":       "C",
    "pit":        "D",
    "sheetpile":  "E",
    "crane":      "F",
    "wp_horiz":   "G",
    "wp_vert":    "H",
    "backfill":   "I",
    "ins_below":  "J",
    # Зона Б1 «Конструктив»
    "plate":      "K",
    "wall":       "L",
    "ceiling":    "M",
    "stair":      "N",
    "ramp":       "O",
    # Зона В «По корпусу»
    "roof":       "P",
    "windows":    "Q",
    "stained":    "R",
    "facade":     "S",
    "lifts":      "T",
    "metal":      "U",
    "doors":      "V",
    "frame":      "W",
    # Зона Б2 «Отделка/инж.»
    "partitions": "X",
    "screed":     "Y",
    "heating":    "Z",
    "water":      "AA",
    "sewage":     "AB",
    "vent":       "AC",
    "elec":       "AD",
    "weak":       "AE",
    "finish":     "AF",
}

# Имя колонки A с лейблом корпуса, B с лейблом этажа
COL_KORPUS = "A"
COL_FLOOR  = "B"


# ── вспомогательные ───────────────────────────────────────────────────────
def is_done(val) -> bool:
    if val is None:
        return False
    if isinstance(val, str):
        v = val.strip()
        if not v or v == "-":
            return False
        return True
    return True


def resolve_building(cfg: dict, key: str) -> dict:
    b = cfg["buildings"][key]
    if "extends" in b:
        tpl = cfg["templates"][b["extends"]]
        merged = dict(tpl)
        for k, v in b.items():
            if k != "extends":
                merged[k] = v
        return merged
    return b


# ── маппинг строки X → (этаж/субблок, col_key) ────────────────────────────
RE_PP_N = re.compile(r"\bпп\s*(\d+)\b", re.IGNORECASE)
RE_ELEV = re.compile(r"([+\-−]?\s*\d+[.,]\d{2,3})")
RE_STEN_N_ETAJ = re.compile(r"стен[аы]?\s*(\d+)[\s\-]*(?:[оеы]?го|ого)?\s*этажа", re.IGNORECASE)
RE_RAMPA = re.compile(r"плит[аы]\s*рампы\s*(ПР\d-?\d?)", re.IGNORECASE)
RE_LESTN_LK = re.compile(r"лестниц(?:ы|ей)?\s+(ЛК\d|ЛМ\d-?\d?)", re.IGNORECASE)
RE_LESTN_PL = re.compile(r"лестничн\w+\s+площадок", re.IGNORECASE)
RE_KROVL = re.compile(r"\bкровл", re.IGNORECASE)


def parse_pp_floor(x: str, b_resolved: dict) -> Optional[str]:
    floors = [f["name"] for f in b_resolved["floors"]]
    m = RE_PP_N.search(x)
    if not m:
        return None
    pp_n = int(m.group(1))
    if pp_n == 0:
        return "Цоколь"
    elev = None
    em = RE_ELEV.search(x.split("ПП")[-1] if "ПП" in x.upper() else x)
    if em:
        try:
            elev = float(em.group(1).replace(",", ".").replace("−", "-").replace(" ", ""))
        except ValueError:
            pass
    if pp_n == 2 and elev is not None and abs(elev - 9.8) < 0.1:
        return "3 этаж" if "3 этаж" in floors else None
    if pp_n == 2 and elev is not None and abs(elev - 6.5) < 0.5:
        return "2 этаж" if "2 этаж" in floors else None
    target = f"{pp_n} этаж"
    return target if target in floors else None


def map_x_to_cell(x: str, d: str, e: str, b_resolved: dict, b_key: str) -> Optional[tuple[str, str]]:
    """(имя_строки, col_key). имя_строки: имя этажа или 'stair:<full>' или 'ramp:<full>'."""
    x_low = (x or "").lower()
    d_low = (d or "").lower() if d else ""
    e_low = (e or "").lower() if e else ""

    if "фундамент" in x_low:
        return ("Цоколь", "plate")

    if "монолит ниже 0" in d_low and "монолитные стены" in e_low:
        return ("Цоколь", "wall")

    is_perekr = ("перекрытие" in x_low) or ("перекрытия" in (d_low + " " + e_low)) or RE_PP_N.search(x or "")

    if "плит" in x_low and "покрыти" in x_low:
        em = RE_ELEV.search(x)
        if em:
            try:
                elev = float(em.group(1).replace(",", ".").replace("−", "-").replace(" ", ""))
            except ValueError:
                elev = None
            if elev is not None:
                if abs(elev - 13.4) < 0.2:
                    return ("4 этаж", "ceiling") if "4 этаж" in [f["name"] for f in b_resolved["floors"]] else None
                if abs(elev - 16.7) < 0.2:
                    return ("Мансарда", "plate") if "Мансарда" in [f["name"] for f in b_resolved["floors"]] else None
        if b_key == "К1":
            return ("2 этаж", "ceiling")
        if b_key == "К2":
            return ("3 этаж", "ceiling")

    if is_perekr and RE_PP_N.search(x or ""):
        f = parse_pp_floor(x, b_resolved)
        if f:
            return (f, "ceiling")

    m = RE_STEN_N_ETAJ.search(x_low) if x else None
    if m:
        n = int(m.group(1))
        floors = [f["name"] for f in b_resolved["floors"]]
        if n == 5 and "Мансарда" in floors:
            return ("Мансарда", "wall")
        target = f"{n} этаж"
        return (target, "wall") if target in floors else None

    m = RE_LESTN_LK.search(x or "")
    if m:
        name = m.group(1)
        for s in b_resolved.get("stairs", []):
            if name.upper() in s.upper():
                return (f"stair:{s}", "stair")
        return None

    if RE_LESTN_PL.search(x or ""):
        for s in b_resolved.get("stairs", []):
            if "подвал" in s.lower():
                return (f"stair:{s}", "stair")
        return None

    m = RE_RAMPA.search(x or "")
    if m:
        code = m.group(1).upper()
        m_num = re.search(r"ПР(\d+)", code)
        if m_num:
            ramp_n = m_num.group(1)
            for r in b_resolved.get("ramps", []):
                if f"ПР{ramp_n}-" in r.upper() or r.startswith(f"ПР{ramp_n}"):
                    return (f"ramp:{r}", "ramp")
        return None

    if RE_KROVL.search(x or ""):
        return ("Кровля", "ceiling")

    return None


# ── сбор: парный Факт ГПР под Зак.ГПР ─────────────────────────────────────
def find_pair_fact_gpr(ws, r_start: int, max_look: int = 4):
    """Ищет AL='Факт ГПР' в r_start+1..r_start+max_look. Возвращает значение AO или None."""
    for off in range(1, max_look + 1):
        if ws.cell(r_start + off, COL_AL).value == "Факт ГПР":
            return ws.cell(r_start + off, COL_AO).value
    return None


def collect_zone_a_below_zero(ws, cfg: dict) -> dict[tuple[str, str, str], int]:
    """Зона А «До 0» — пишется в строку «Цоколь» каждого корпуса.

    Источник готовности — пары Зак.ГПР / Факт ГПР, по разделам МСГ:
      Подгот.    : раздел «ПОДГОТОВИТЕЛЬНЫЕ РАБОТЫ» — AND по 4 подвидам, Y=null → все 6
      Котлован   : Зак.ГПР «Откопка котлована», Y=К1..К6
      Шпунт      : Зак.ГПР «Устройство шпунта», Y=К1/К2/К3 (К4-К6 в шаблоне N/A)
      БК         : Зак.ГПР «БАШЕННЫЙ КРАН», Y='БК' → все 6
      ГИ гор.    : Зак.ГПР «ГИДРОИЗОЛЯЦИЯ … гидрошпонка, инжект-система», Y=К1..К6 + К1-К6
      ГИ верт.   : Зак.ГПР «ГИДРОИЗОЛЯЦИЯ … наружных стен», Y=К1..К6 + К1-К6
      Обр.засыпка: Зак.ГПР «ОБРАТНАЯ ЗАСЫПКА НИЖЕ ОТМ. 0.000», Y=К1..К6 + К1-К6
      Утепл. ниже 0: Зак.ГПР «УТЕПЛЕНИЕ … наружных стен», Y=К1..К6 + К1-К6
    """
    result: dict[tuple[str, str, str], int] = {}

    # Подготовка: 4 подвида, считаем готовым если ВСЕ четыре закрыты (Y отсутствует → один статус)
    prep_subs = (
        "Вырубка деревьев",
        "Монтаж временных дорог",
        "Монтаж временных инженерных сетей",
        "Мобилизация на объекте",
    )
    prep_done_count = 0
    prep_total = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        x = ws.cell(r, COL_X).value or ""
        if not any(sub in x for sub in prep_subs):
            continue
        prep_total += 1
        fact_ao = find_pair_fact_gpr(ws, r)
        if is_done(fact_ao):
            prep_done_count += 1
    prep_val = 1 if (prep_total > 0 and prep_done_count == prep_total) else 0
    for corpus in CORPUSES:
        result[(corpus, "Цоколь", "prep")] = prep_val

    # Простые мапы: «вид X в строке Зак.ГПР» → col_key
    simple_map = {
        "Откопка котлована":       "pit",
        "Устройство шпунта":       "sheetpile",
        "БАШЕННЫЙ КРАН":           "crane",
    }
    bk_val = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        x = ws.cell(r, COL_X).value or ""
        y = ws.cell(r, COL_Y).value
        for needle, col_key in simple_map.items():
            if needle in x:
                fact_ao = find_pair_fact_gpr(ws, r)
                val = 1 if is_done(fact_ao) else 0
                if col_key == "crane":
                    bk_val = val  # один БК на стройку
                elif col_key == "sheetpile" and y in ("К1", "К2", "К3"):
                    result[(y, "Цоколь", col_key)] = val
                elif col_key == "pit" and y in CORPUSES:
                    result[(y, "Цоколь", col_key)] = val
    if bk_val is not None:
        for corpus in CORPUSES:
            result[(corpus, "Цоколь", "crane")] = bk_val

    # Гидроизоляция/утепление/засыпка — 4 подвида, Y разные
    gi_map = {
        "ГИДРОИЗОЛЯЦИЯ НИЖЕ ОТМ. 0.000: гидрошпонка": "wp_horiz",
        "ГИДРОИЗОЛЯЦИЯ НИЖЕ ОТМ. 0.000: наружных стен": "wp_vert",
        "ОБРАТНАЯ ЗАСЫПКА НИЖЕ ОТМ. 0.000": "backfill",
        "УТЕПЛЕНИЕ НИЖЕ ОТМ. 0.000: наружных стен": "ins_below",
    }
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        x = ws.cell(r, COL_X).value or ""
        y = ws.cell(r, COL_Y).value
        for needle, col_key in gi_map.items():
            if x.strip().startswith(needle.rstrip()) or needle in x:
                fact_ao = find_pair_fact_gpr(ws, r)
                val = 1 if is_done(fact_ao) else 0
                targets = []
                if y in CORPUSES:
                    targets = [y]
                elif y == "К1-К6" or y is None:
                    targets = list(CORPUSES)
                for c in targets:
                    key = (c, "Цоколь", col_key)
                    # AND-логика: «1» только если все источники = 1; «0» если хоть где «0»
                    if key in result:
                        if val == 0:
                            result[key] = 0
                    else:
                        result[key] = val

    return result


def collect_zone_v_by_building(ws, cfg: dict) -> dict[tuple[str, str, str], int]:
    """Зона В «По корпусу» — пишется в строку «По корпусу в целом» каждого корпуса.

    Все 8 видов — через парный Факт ГПР под Зак.ГПР.

      Кровля    : «Устройство (металлической|эксплуатируемой) кровли», Y=К1..К6 (AND если 2 подвида)
      Окна Al   : «Монтаж оконных и балконных блоков», Y=К1, К3..К6
      Витражи   : «Монтаж витражных конструкций», Y=ИК+К1..К6 (ИК → все 6)
      Фасад     : «Устройство фасадов», Y=К1..К6
      Лифты     : «Монтаж лифтового оборудования», Y=К1, К3..К6, К1-К6
      МК        : «Монтаж металлических конструкций», Y=К1..К6, К1-К6
      Двери     : «Монтаж дверных блоков», Y=К1..К6, К1-К6
      Сборн.каркас : «Монтаж сборных элементов каркаса», Y=К3..К6
    """
    result: dict[tuple[str, str, str], int] = {}

    work_map = [
        # (needle_in_X, col_key, applies_corpuses)
        ("Устройство металлической кровли",       "roof",    None),
        ("Устройство эксплуатируемой кровли",     "roof",    None),
        ("Монтаж оконных и балконных блоков",     "windows", None),
        ("Монтаж витражных конструкций",          "stained", None),
        ("Устройство фасадов",                    "facade",  None),
        ("Монтаж лифтового оборудования",         "lifts",   None),
        ("Монтаж металлических конструкций",      "metal",   None),
        ("Монтаж дверных блоков",                 "doors",   None),
        ("Монтаж сборных элементов каркаса",      "frame",   ("К3", "К4", "К5", "К6")),
    ]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        x = ws.cell(r, COL_X).value or ""
        y = ws.cell(r, COL_Y).value
        for needle, col_key, restricted in work_map:
            if needle in x:
                fact_ao = find_pair_fact_gpr(ws, r)
                val = 1 if is_done(fact_ao) else 0
                if y in CORPUSES:
                    targets = [y]
                elif y in ("К1-К6", "ИК", None, "БК"):
                    targets = list(CORPUSES)
                else:
                    targets = []
                if restricted:
                    targets = [c for c in targets if c in restricted]
                for c in targets:
                    key = (c, "По корпусу в целом", col_key)
                    # AND-логика (важно для кровли — 2 подвида)
                    if key in result:
                        if val == 0:
                            result[key] = 0
                    else:
                        result[key] = val

    return result


def collect_object_works(ws, cfg: dict) -> dict[tuple[str, str, str], int]:
    """Общестроит (ИТП/Благоустройство/Нар.сети/Ввод) — по 1 значению на «объект».

    Считаем готовым раздел если ВСЕ Зак.ГПР строки раздела закрыты.
    Ключ строки шаблона — название (без «- Мещеряков» и т.п. суффиксов).
    """
    result: dict[tuple[str, str, str], int] = {}

    # Разделы и их «короткие» имена в шаблоне
    sections = [
        ("Монтаж ИТП",                   "ИТП"),
        ("Благоустройство",              "Благоустройство"),
        ("Наружные инженерные сети",     "Наружные инженерные сети"),
        ("ВВОД ОБЪЕКТА В ЭКСПЛУАТАЦИЮ",  "Ввод объекта в эксплуатацию"),
    ]
    # Сначала найдём для каждого раздела диапазон строк (от section header до следующего)
    section_headers = []
    for r in range(2, ws.max_row + 1):
        x = ws.cell(r, COL_X).value
        al = ws.cell(r, COL_AL).value
        if x and not al:
            section_headers.append((r, str(x).strip()))
    section_headers.append((ws.max_row + 1, "END"))

    for i in range(len(section_headers) - 1):
        r_start, name = section_headers[i]
        r_end, _ = section_headers[i + 1]
        for src_prefix, tpl_label in sections:
            if name.startswith(src_prefix):
                # Считаем все Зак.ГПР внутри
                total = 0
                done = 0
                for r in range(r_start + 1, r_end):
                    if ws.cell(r, COL_AL).value == "Зак. ГПР":
                        total += 1
                        if is_done(find_pair_fact_gpr(ws, r)):
                            done += 1
                val = 1 if (total > 0 and done == total) else 0
                # ключ — особый: corpus=None, row_name = "obj:<tpl_label>", col_key="total"
                result[("OBJ", f"obj:{tpl_label}", "total")] = val
                break

    return result


# ── сбор готовности из gsheet (зона Б1 + Кровля) ──────────────────────────
def collect_done(gsheet_path: Path, cfg: dict) -> dict[tuple[str, str, str], int]:
    wb = load_workbook(gsheet_path, data_only=True)
    target = next((n for n in wb.sheetnames if n.startswith("МСГ")), wb.sheetnames[0])
    ws = wb[target]

    resolved = {k: resolve_building(cfg, k) for k in cfg["buildings"].keys()}
    result: dict[tuple[str, str, str], int] = {}

    # 1) Бетонирование — план/факт
    for r in range(2, ws.max_row + 1):
        x = ws.cell(r, COL_X).value
        if not x or not str(x).lower().startswith("бетонирование"):
            continue
        y = ws.cell(r, COL_Y).value
        if y not in CORPUSES:
            continue
        if ws.cell(r, COL_AL).value != "План":
            continue
        if ws.cell(r + 1, COL_AL).value != "Факт":
            continue
        fact_ao = ws.cell(r + 1, COL_AO).value
        d = ws.cell(r, COL_D).value
        e = ws.cell(r, COL_E).value
        cell = map_x_to_cell(x, d, e, resolved[y], y)
        if cell is None:
            continue
        floor_name, col_key = cell
        key = (y, floor_name, col_key)
        val = 1 if is_done(fact_ao) else 0
        if key in result:
            if val == 1:
                result[key] = 1
        else:
            result[key] = val

    # 2) Кровля — через Зак.ГПР / Факт ГПР
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_AL).value != "Зак. ГПР":
            continue
        y = ws.cell(r, COL_Y).value
        if y not in CORPUSES:
            continue
        x = ws.cell(r, COL_X).value or ""
        if "кровл" not in x.lower() or "устройств" not in x.lower():
            continue
        fact_ao = None
        for off in (1, 2, 3, 4):
            if ws.cell(r + off, COL_AL).value == "Факт ГПР":
                fact_ao = ws.cell(r + off, COL_AO).value
                break
        key = (y, "Кровля", "ceiling")
        val = 1 if is_done(fact_ao) else 0
        if key in result:
            if val == 1:
                result[key] = 1
        else:
            result[key] = val

    return result


# ── поиск строки в шаблоне ────────────────────────────────────────────────
def find_target_row(ws_t, b_key: str, row_name: str) -> Optional[int]:
    """
    Найти строку шаблона по корпусу + имени.
    row_name либо имя этажа («Цоколь», «1 этаж», ...), либо «stair:<full>», либо «ramp:<full>».
    Идём блоком от заголовка «Корпус Кi» (в A, merged), читаем B-колонку, ищем совпадение по префиксу.
    """
    if row_name.startswith("stair:"):
        kind, want = "stair", row_name[len("stair:"):]
    elif row_name.startswith("ramp:"):
        kind, want = "ramp", row_name[len("ramp:"):]
    elif row_name == "По корпусу в целом":
        kind, want = "bsummary", ""
    else:
        kind, want = "floor", row_name

    in_section = False
    for r in range(1, ws_t.max_row + 1):
        a = ws_t.cell(r, 1).value
        b = ws_t.cell(r, 2).value

        # начало секции
        if a and isinstance(a, str) and a.strip() == f"Корпус {b_key}":
            in_section = True
            continue
        # выход из секции
        if in_section and a and isinstance(a, str):
            a_str = a.strip()
            if a_str.startswith("Корпус ") and a_str != f"Корпус {b_key}":
                return None

        if not in_section:
            continue

        if not b or not isinstance(b, str):
            continue
        b_str = b.strip()

        if b_str.startswith("Итог по корпусу"):
            return None

        if kind == "bsummary":
            if b_str == "По корпусу в целом":
                return r
            continue

        if b_str == "По корпусу в целом":
            continue

        if kind == "floor":
            # имя этажа может быть «Цоколь   (отм. ...)» — отрезаем хвост
            head = b_str.split("(отм.")[0].strip() if "(отм." in b_str else b_str
            if head == want:
                return r
        elif kind == "stair":
            if b_str.startswith("Лестница:"):
                stair_name = b_str[len("Лестница:"):].strip()
                if stair_name == want:
                    return r
        elif kind == "ramp":
            if b_str.startswith("Рампа:"):
                ramp_name = b_str[len("Рампа:"):].strip()
                if ramp_name == want:
                    return r

    return None


NA_GRAY = "D9D9D9"


def is_na_cell(cell) -> bool:
    """N/A определяется по серой заливке шаблона (C_NA_BG = 'D9D9D9')."""
    try:
        rgb = cell.fill.fgColor.rgb
        if isinstance(rgb, str) and rgb.upper().endswith(NA_GRAY):
            return True
    except Exception:
        pass
    return False


def is_writable(cell) -> bool:
    return not is_na_cell(cell)


# Колонки Зоны Б2 «Отделка/инж.» — для дефолтного заполнения нулями
ZONE_B2_COLS = ("X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", type=Path, default=CONFIG_PATH)
    ap.add_argument("--gsheet", type=Path, default=None)
    ap.add_argument("--template", type=Path, default=None)
    args = ap.parse_args()

    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    with args.config.open(encoding="utf-8") as f:
        cfg = json.load(f)

    if args.template:
        tpl_path = args.template
    else:
        tpl_path = Path(cfg["output"]["dir"]) / DEFAULT_TEMPLATE_NAME

    # 1) Скачать gsheet
    if args.gsheet is None:
        gid = cfg["source"]["gdrive_file_id"]
        tmp = Path(tempfile.gettempdir()) / f"repino_msg_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        downloader = Path(r"C:\Авраменко\Claude Code Projects\МСГ\scripts\gsheet_download_xlsx.py")
        print(f"[1] Скачиваю gsheet -> {tmp}")
        subprocess.check_call([sys.executable, str(downloader), "--file-id", gid, "--out", str(tmp)])
        gsheet_path = tmp
    else:
        gsheet_path = args.gsheet
        print(f"[1] Используем переданный gsheet: {gsheet_path}")

    # 2) Собираем карту готовности
    print(f"[2] Читаю «МСГ, ГПР»...")
    wb_src = load_workbook(gsheet_path, data_only=True)
    src_sheet = next((n for n in wb_src.sheetnames if n.startswith("МСГ")), wb_src.sheetnames[0])
    ws_src = wb_src[src_sheet]

    done_map = collect_done(gsheet_path, cfg)
    print(f"  Б1+Кровля: {len(done_map)} ячеек")

    zone_a = collect_zone_a_below_zero(ws_src, cfg)
    print(f"  Зона А «До 0»: {len(zone_a)} ячеек")
    done_map.update(zone_a)

    zone_v = collect_zone_v_by_building(ws_src, cfg)
    print(f"  Зона В «По корпусу»: {len(zone_v)} ячеек")
    done_map.update(zone_v)

    obj_works = collect_object_works(ws_src, cfg)
    print(f"  Общестроит: {len(obj_works)} ячеек")
    done_map.update(obj_works)

    n_ones = sum(1 for v in done_map.values() if v == 1)
    print(f"  ИТОГО целевых: {len(done_map)}  | =1: {n_ones}  | =0: {len(done_map) - n_ones}")

    # 3) Открываем шаблон, пишем
    print(f"[3] Пишу в шаблон: {tpl_path}")
    wb = load_workbook(tpl_path)
    ws = wb["Таблица 1-0"]

    n_written = 0
    n_skipped_na = 0
    n_missing = 0
    skipped: list = []
    for (corpus, row_name, col_key), val in sorted(done_map.items()):
        # Общестроительные: запись в колонку AG строки obj:<label>
        if corpus == "OBJ" and row_name.startswith("obj:"):
            label = row_name[len("obj:"):]
            target_r = None
            for r in range(1, ws.max_row + 1):
                b = ws.cell(r, 2).value
                if b and str(b).strip() == label:
                    target_r = r
                    break
            if target_r is None:
                n_missing += 1
                skipped.append((corpus, row_name, col_key, val, "obj row not found"))
                continue
            cell = ws.cell(target_r, 33)  # AG
            cell.value = val
            n_written += 1
            continue

        row = find_target_row(ws, corpus, row_name)
        if row is None:
            n_missing += 1
            skipped.append((corpus, row_name, col_key, val, "row not found"))
            continue
        col_letter = TPL_COLS.get(col_key)
        if col_letter is None:
            n_missing += 1
            skipped.append((corpus, row_name, col_key, val, "col key not in TPL_COLS"))
            continue
        cell = ws[f"{col_letter}{row}"]
        if not is_writable(cell):
            n_skipped_na += 1
            skipped.append((corpus, row_name, col_key, val, "cell is N/A"))
            continue
        cell.value = val
        n_written += 1

    print(f"    Записано: {n_written}  | N/A пропущено: {n_skipped_na}  | не сопоставлено: {n_missing}")
    if skipped:
        print(f"    Пропуски (первые 20):")
        for s in skipped[:20]:
            print(f"      {s}")

    # 4) Дефолт 0 для активных пустых ячеек Зоны Б2 (Отделка/инж.) —
    #    источника готовности в МСГ по этажам нет; пользователь переключит вручную.
    from openpyxl.cell.cell import MergedCell
    n_zone_b2_defaults = 0
    for r in range(5, ws.max_row + 1):
        for col_letter in ZONE_B2_COLS:
            cell = ws[f"{col_letter}{r}"]
            if isinstance(cell, MergedCell):
                continue
            if is_na_cell(cell):
                continue
            if cell.value is None:
                cell.value = 0
                n_zone_b2_defaults += 1
    print(f"    Зона Б2 — дефолт 0 в активных пустых ячейках: {n_zone_b2_defaults}")

    # 5) Сохранить
    wb.save(tpl_path)
    print(f"[OK] -> {tpl_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
