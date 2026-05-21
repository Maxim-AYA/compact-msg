"""
build_1_0_template.py — генератор пустого шаблона «Таблица 1-0 Репино.xlsx».

Читает config.json в родительской папке, строит один лист со всеми корпусами К1..К6:
- двухуровневая шапка
- блоки корпусов (заголовок-разделитель + строки-этажи + блок Лестницы + блок Рампы)
- итоги строк (% по 5 колонкам данных)
- итоги корпусов и общий итог по объекту
- условное форматирование <50/50-99/100 на ячейках итогов
- легенда «1 = готово, 0 = не готово, пусто = не требуется»

Использование:
    python build_1_0_template.py
    python build_1_0_template.py --out "path\\to\\custom.xlsx"
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CONFIG_PATH = ROOT / "config.json"


# ── палитра ────────────────────────────────────────────────────────────────
C_TITLE_BG = "1F4E79"   # тёмно-синий
C_TITLE_FG = "FFFFFF"
C_HEADER_BG = "2E75B6"  # синий шапки
C_HEADER_FG = "FFFFFF"
C_GROUP_BG = "D9E1F2"   # бледно-синий (название корпуса)
C_SUBGROUP_BG = "E7E6E6"  # серый (Лестницы / Рампы)
C_FLOOR_BG = "F2F2F2"   # очень бледный серый (зебра этажей)
C_TOTAL_BG = "BDD7EE"   # средне-синий (итоги)
C_RED = "F8CBAD"
C_YELLOW = "FFE699"
C_GREEN = "C6EFCE"
# дискретный CF для ячеек данных (как в «Поэтажном плане»)
C_CELL_DONE_BG = "E2EFD9"   # бледно-зелёный (значение 1)
C_CELL_DONE_FG = "00B050"   # тёмно-зелёный (текст для 1)
C_CELL_NOT_BG = "FBE4D5"    # бледно-персиковый (значение 0)
C_GRID = "BFBFBF"

FONT_BASE = "Calibri"


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def thin_border() -> Border:
    s = Side(style="thin", color=C_GRID)
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border() -> Border:
    s = Side(style="medium", color="404040")
    return Border(left=s, right=s, top=s, bottom=s)


# ── конфиг ─────────────────────────────────────────────────────────────────
def load_config(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def resolve_building(cfg: dict, key: str) -> dict:
    b = cfg["buildings"][key]
    if "extends" in b:
        tpl_key = b["extends"]
        tpl = cfg["templates"][tpl_key]
        merged = dict(tpl)
        for k, v in b.items():
            if k != "extends":
                merged[k] = v
        return merged
    return b


# ── расчёт строк ───────────────────────────────────────────────────────────
def floor_data_keys(floor: dict, columns: list[dict]) -> list[str]:
    """Какие колонки заполняются на этой строке-этаже (только Плита/Стены/Перекрытие)."""
    result = []
    for col in columns:
        if col["key"] in ("stair", "ramp"):
            continue  # на этажных строках лестницы/рампы пустые
        val = floor.get(col["key"])
        if val:
            result.append(col["key"])
    return result


# ── основной билдер ────────────────────────────────────────────────────────
def build(cfg: dict, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица 1-0"

    columns: list[dict] = cfg["columns"]
    n_data_cols = len(columns)

    # Раскладка колонок:
    # A — № / Корпус·Этаж (название)
    # B..(1+n_data_cols) — данные конструктивов
    # последняя — Итог по строке (%)
    COL_NAME = 1
    COL_DATA_FIRST = 2
    COL_DATA_LAST = COL_DATA_FIRST + n_data_cols - 1
    COL_TOTAL = COL_DATA_LAST + 1
    TOTAL_COLS = COL_TOTAL

    # ── строки 1-3: заголовок + легенда ──
    ws.cell(1, 1, value=f"Таблица 1-0 — {cfg['project_full_name']}").font = Font(
        name=FONT_BASE, size=16, bold=True, color=C_TITLE_FG
    )
    ws.cell(1, 1).fill = fill(C_TITLE_BG)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    ws.row_dimensions[1].height = 28

    legend = "Легенда:   1 = готово   ·   0 = не готово   ·   пусто = не требуется"
    ws.cell(2, 1, value=legend).font = Font(name=FONT_BASE, size=10, italic=True, color="595959")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    ws.row_dimensions[2].height = 18

    # ── строки 3-4: двухуровневая шапка ──
    HEADER_ROW_1 = 3
    HEADER_ROW_2 = 4

    # верхняя группа: «Корпус / Этаж» (A merged 2 строки)
    c = ws.cell(HEADER_ROW_1, COL_NAME, value="Корпус / Этаж / Конструктив")
    ws.merge_cells(start_row=HEADER_ROW_1, start_column=COL_NAME, end_row=HEADER_ROW_2, end_column=COL_NAME)

    # верхняя группа: разделение data-колонок по семействам (Конструктивы этажа vs Лестницы/Рампы)
    # ищем индексы plate/wall/ceiling и stair/ramp
    structurals = [i for i, col in enumerate(columns) if col["key"] in ("plate", "wall", "ceiling")]
    sr = [i for i, col in enumerate(columns) if col["key"] in ("stair", "ramp")]
    if structurals:
        s_first = COL_DATA_FIRST + structurals[0]
        s_last = COL_DATA_FIRST + structurals[-1]
        ws.cell(HEADER_ROW_1, s_first, value="Конструктивы этажа")
        if s_first != s_last:
            ws.merge_cells(start_row=HEADER_ROW_1, start_column=s_first, end_row=HEADER_ROW_1, end_column=s_last)
    if sr:
        r_first = COL_DATA_FIRST + sr[0]
        r_last = COL_DATA_FIRST + sr[-1]
        ws.cell(HEADER_ROW_1, r_first, value="Лестницы / Рампы")
        if r_first != r_last:
            ws.merge_cells(start_row=HEADER_ROW_1, start_column=r_first, end_row=HEADER_ROW_1, end_column=r_last)

    # верхняя группа: Итог (merged 2 строки)
    ws.cell(HEADER_ROW_1, COL_TOTAL, value="Итог")
    ws.merge_cells(start_row=HEADER_ROW_1, start_column=COL_TOTAL, end_row=HEADER_ROW_2, end_column=COL_TOTAL)

    # нижняя строка шапки: имена 5 колонок
    for i, col in enumerate(columns):
        ws.cell(HEADER_ROW_2, COL_DATA_FIRST + i, value=col["title"])

    # стиль шапки
    for r in (HEADER_ROW_1, HEADER_ROW_2):
        for c_idx in range(1, TOTAL_COLS + 1):
            cell = ws.cell(r, c_idx)
            cell.fill = fill(C_HEADER_BG)
            cell.font = Font(name=FONT_BASE, size=11, bold=True, color=C_HEADER_FG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
    ws.row_dimensions[HEADER_ROW_1].height = 22
    ws.row_dimensions[HEADER_ROW_2].height = 30

    # ── блоки корпусов ──
    row = HEADER_ROW_2 + 1
    building_total_rows: list[tuple[str, int]] = []  # для общего итога

    for b_key in cfg["buildings"].keys():
        b = resolve_building(cfg, b_key)

        # Заголовок корпуса (merged по всем колонкам)
        title_row = row
        cell = ws.cell(row, COL_NAME, value=f"Корпус {b_key}")
        ws.merge_cells(start_row=row, start_column=COL_NAME, end_row=row, end_column=TOTAL_COLS)
        cell.fill = fill(C_GROUP_BG)
        cell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = thin_border()
        ws.row_dimensions[row].height = 22
        row += 1

        # Этажи
        floor_rows: list[int] = []
        for floor in b["floors"]:
            r = row
            floor_rows.append(r)

            label = f"  {floor['name']}"
            if floor.get("elev"):
                label += f"  (отм. {floor['elev']})"
            name_cell = ws.cell(r, COL_NAME, value=label)
            name_cell.font = Font(name=FONT_BASE, size=10)
            name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            name_cell.border = thin_border()

            for i, col in enumerate(columns):
                c_idx = COL_DATA_FIRST + i
                cell = ws.cell(r, c_idx)
                cell.border = thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=FONT_BASE, size=10)
                # Серая заливка для «не применимо» убрана — пустые ячейки остаются прозрачными.

            # Итог строки этажа
            data_range = f"{get_column_letter(COL_DATA_FIRST)}{r}:{get_column_letter(COL_DATA_LAST)}{r}"
            tcell = ws.cell(r, COL_TOTAL,
                            value=f'=IF(COUNT({data_range})=0,"",SUM({data_range})/COUNT({data_range}))')
            tcell.number_format = "0%"
            tcell.alignment = Alignment(horizontal="center", vertical="center")
            tcell.font = Font(name=FONT_BASE, size=10, bold=True)
            tcell.border = thin_border()

            # лёгкая зебра
            if (r - HEADER_ROW_2) % 2 == 0:
                for c_idx in range(COL_NAME, TOTAL_COLS):
                    if ws.cell(r, c_idx).fill.fgColor.rgb in ("00000000", None):
                        ws.cell(r, c_idx).fill = fill(C_FLOOR_BG)

            row += 1

        # Блок «Лестницы» (если есть)
        stair_rows: list[int] = []
        if b.get("stairs"):
            # подзаголовок
            sub = ws.cell(row, COL_NAME, value="  Лестницы")
            ws.merge_cells(start_row=row, start_column=COL_NAME, end_row=row, end_column=TOTAL_COLS)
            sub.fill = fill(C_SUBGROUP_BG)
            sub.font = Font(name=FONT_BASE, size=10, bold=True, italic=True, color="595959")
            sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            sub.border = thin_border()
            row += 1

            for stair_name in b["stairs"]:
                r = row
                stair_rows.append(r)
                name_cell = ws.cell(r, COL_NAME, value=f"    {stair_name}")
                name_cell.font = Font(name=FONT_BASE, size=10)
                name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
                name_cell.border = thin_border()
                for i, col in enumerate(columns):
                    c_idx = COL_DATA_FIRST + i
                    cell = ws.cell(r, c_idx)
                    cell.border = thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name=FONT_BASE, size=10)
                # Итог: тут будет только 1 заполненная ячейка из 5
                data_range = f"{get_column_letter(COL_DATA_FIRST)}{r}:{get_column_letter(COL_DATA_LAST)}{r}"
                tcell = ws.cell(r, COL_TOTAL,
                                value=f'=IF(COUNT({data_range})=0,"",SUM({data_range})/COUNT({data_range}))')
                tcell.number_format = "0%"
                tcell.alignment = Alignment(horizontal="center", vertical="center")
                tcell.font = Font(name=FONT_BASE, size=10, bold=True)
                tcell.border = thin_border()
                row += 1

        # Блок «Рампы» (если есть)
        ramp_rows: list[int] = []
        if b.get("ramps"):
            sub = ws.cell(row, COL_NAME, value="  Рампы")
            ws.merge_cells(start_row=row, start_column=COL_NAME, end_row=row, end_column=TOTAL_COLS)
            sub.fill = fill(C_SUBGROUP_BG)
            sub.font = Font(name=FONT_BASE, size=10, bold=True, italic=True, color="595959")
            sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            sub.border = thin_border()
            row += 1

            for ramp_name in b["ramps"]:
                r = row
                ramp_rows.append(r)
                name_cell = ws.cell(r, COL_NAME, value=f"    {ramp_name}")
                name_cell.font = Font(name=FONT_BASE, size=10)
                name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
                name_cell.border = thin_border()
                for i, col in enumerate(columns):
                    c_idx = COL_DATA_FIRST + i
                    cell = ws.cell(r, c_idx)
                    cell.border = thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name=FONT_BASE, size=10)
                data_range = f"{get_column_letter(COL_DATA_FIRST)}{r}:{get_column_letter(COL_DATA_LAST)}{r}"
                tcell = ws.cell(r, COL_TOTAL,
                                value=f'=IF(COUNT({data_range})=0,"",SUM({data_range})/COUNT({data_range}))')
                tcell.number_format = "0%"
                tcell.alignment = Alignment(horizontal="center", vertical="center")
                tcell.font = Font(name=FONT_BASE, size=10, bold=True)
                tcell.border = thin_border()
                row += 1

        # Итог по корпусу — диапазон всех data-строк блока (этажи + лестницы + рампы)
        all_data_rows = floor_rows + stair_rows + ramp_rows
        if all_data_rows:
            total_row = row
            building_total_rows.append((b_key, total_row))

            name_cell = ws.cell(total_row, COL_NAME, value=f"Итог по корпусу {b_key}")
            name_cell.font = Font(name=FONT_BASE, size=11, bold=True, color="1F4E79")
            name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            name_cell.fill = fill(C_TOTAL_BG)
            name_cell.border = medium_border()

            # Итог по каждой data-колонке
            r_first = min(all_data_rows)
            r_last = max(all_data_rows)
            for i, col in enumerate(columns):
                c_idx = COL_DATA_FIRST + i
                col_letter = get_column_letter(c_idx)
                cell = ws.cell(total_row, c_idx,
                               value=f'=IF(COUNT({col_letter}{r_first}:{col_letter}{r_last})=0,"",'
                                     f'SUM({col_letter}{r_first}:{col_letter}{r_last})/'
                                     f'COUNT({col_letter}{r_first}:{col_letter}{r_last}))')
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=FONT_BASE, size=11, bold=True)
                cell.fill = fill(C_TOTAL_BG)
                cell.border = medium_border()

            # Общий итог корпуса (среднее по всем заполненным)
            grand_range = f"{get_column_letter(COL_DATA_FIRST)}{r_first}:{get_column_letter(COL_DATA_LAST)}{r_last}"
            tcell = ws.cell(total_row, COL_TOTAL,
                            value=f'=IF(COUNT({grand_range})=0,"",SUM({grand_range})/COUNT({grand_range}))')
            tcell.number_format = "0%"
            tcell.alignment = Alignment(horizontal="center", vertical="center")
            tcell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
            tcell.fill = fill(C_TOTAL_BG)
            tcell.border = medium_border()
            ws.row_dimensions[total_row].height = 22
            row += 1

        # Пустая строка между корпусами
        ws.row_dimensions[row].height = 6
        row += 1

    # ── Общий итог по объекту ──
    grand_row = row
    name_cell = ws.cell(grand_row, COL_NAME, value="ИТОГО по объекту Репино")
    name_cell.font = Font(name=FONT_BASE, size=12, bold=True, color="FFFFFF")
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = fill(C_TITLE_BG)
    name_cell.border = medium_border()

    # Среднее по всем итогам корпусов (с весом 1:1 — у пользователя нет другого ТЗ)
    for i, col in enumerate(columns):
        c_idx = COL_DATA_FIRST + i
        col_letter = get_column_letter(c_idx)
        refs = ",".join(f"{col_letter}{r}" for _, r in building_total_rows)
        cell = ws.cell(grand_row, c_idx, value=f'=IFERROR(AVERAGE({refs}),"")')
        cell.number_format = "0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BASE, size=12, bold=True, color="FFFFFF")
        cell.fill = fill(C_TITLE_BG)
        cell.border = medium_border()

    total_refs = ",".join(f"{get_column_letter(COL_TOTAL)}{r}" for _, r in building_total_rows)
    tcell = ws.cell(grand_row, COL_TOTAL, value=f'=IFERROR(AVERAGE({total_refs}),"")')
    tcell.number_format = "0%"
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.font = Font(name=FONT_BASE, size=14, bold=True, color="FFFFFF")
    tcell.fill = fill(C_TITLE_BG)
    tcell.border = medium_border()
    ws.row_dimensions[grand_row].height = 28
    row += 1

    # ── ширины колонок ──
    ws.column_dimensions[get_column_letter(COL_NAME)].width = 42
    for i in range(n_data_cols):
        col_letter = get_column_letter(COL_DATA_FIRST + i)
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 12

    # ── дискретный CF на data-ячейках (1 = зелёный, 0 = персиковый) ──
    # Применяем ко всем data-колонкам всех data-строк (этажи + лестницы + рампы).
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    dxf_done = DifferentialStyle(
        font=Font(color=C_CELL_DONE_FG, bold=True),
        fill=fill(C_CELL_DONE_BG),
    )
    dxf_not = DifferentialStyle(
        fill=fill(C_CELL_NOT_BG),
    )
    data_cf_range = (
        f"{get_column_letter(COL_DATA_FIRST)}{HEADER_ROW_2 + 1}:"
        f"{get_column_letter(COL_DATA_LAST)}{grand_row - 1}"
    )
    ws.conditional_formatting.add(
        data_cf_range,
        Rule(type="cellIs", operator="equal", formula=["1"], dxf=dxf_done, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        data_cf_range,
        Rule(type="cellIs", operator="equal", formula=["0"], dxf=dxf_not, stopIfTrue=False),
    )

    # ── условное форматирование на колонке итогов (процентная шкала) ──
    last_row = grand_row
    cf_range = f"{get_column_letter(COL_TOTAL)}{HEADER_ROW_2 + 1}:{get_column_letter(COL_TOTAL)}{last_row}"
    # Применяем также к итогам по каждой data-колонке у строк-итогов корпусов и общего итога
    th = cfg.get("thresholds_pct", {})
    red_below = th.get("red_below", 50) / 100.0
    yellow_below = th.get("yellow_below", 100) / 100.0

    rules = [
        (CellIsRule(operator="lessThan", formula=[str(red_below)],
                    stopIfTrue=False, fill=fill(C_RED)), "red"),
        (CellIsRule(operator="lessThan", formula=[str(yellow_below)],
                    stopIfTrue=False, fill=fill(C_YELLOW)), "yellow"),
        (CellIsRule(operator="greaterThanOrEqual", formula=[str(yellow_below)],
                    stopIfTrue=False, fill=fill(C_GREEN)), "green"),
    ]
    for rule, _ in rules:
        ws.conditional_formatting.add(cf_range, rule)

    # Также добавим CF на итоговые строки корпусов и общий итог по data-колонкам
    for _, tr in building_total_rows:
        rng = f"{get_column_letter(COL_DATA_FIRST)}{tr}:{get_column_letter(COL_DATA_LAST)}{tr}"
        for rule, _label in [
            (CellIsRule(operator="lessThan", formula=[str(red_below)], fill=fill(C_RED)), "r"),
            (CellIsRule(operator="lessThan", formula=[str(yellow_below)], fill=fill(C_YELLOW)), "y"),
            (CellIsRule(operator="greaterThanOrEqual", formula=[str(yellow_below)], fill=fill(C_GREEN)), "g"),
        ]:
            ws.conditional_formatting.add(rng, rule)

    grand_rng = f"{get_column_letter(COL_DATA_FIRST)}{grand_row}:{get_column_letter(COL_DATA_LAST)}{grand_row}"
    for rule, _label in [
        (CellIsRule(operator="lessThan", formula=[str(red_below)], fill=fill(C_RED)), "r"),
        (CellIsRule(operator="lessThan", formula=[str(yellow_below)], fill=fill(C_YELLOW)), "y"),
        (CellIsRule(operator="greaterThanOrEqual", formula=[str(yellow_below)], fill=fill(C_GREEN)), "g"),
    ]:
        ws.conditional_formatting.add(grand_rng, rule)

    # ── фиксация шапки ──
    ws.freeze_panes = f"{get_column_letter(COL_DATA_FIRST)}{HEADER_ROW_2 + 1}"

    # ── печать: альбомная, ширина — лист ──
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    # ── расшифровки в виде «комментариев» (через rowDescriptions / Hint) ──
    # openpyxl не поддерживает tooltips нативно. Положу справку конструктивов на отдельный лист.
    info_ws = wb.create_sheet("Справка по конструктивам")
    info_ws.cell(1, 1, value="Расшифровка конструктивов по корпусам (источник — раздел РД 6.3.8 КЖ, КЖИ)").font = Font(
        bold=True, size=12
    )
    info_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    info_ws.cell(2, 1, value="Корпус")
    info_ws.cell(2, 2, value="Этаж")
    info_ws.cell(2, 3, value="Плита")
    info_ws.cell(2, 4, value="Стены")
    info_ws.cell(2, 5, value="Перекрытие")
    info_ws.cell(2, 6, value="Примечание")
    for c_idx in range(1, 7):
        cc = info_ws.cell(2, c_idx)
        cc.font = Font(bold=True, color=C_HEADER_FG)
        cc.fill = fill(C_HEADER_BG)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = thin_border()

    ir = 3
    for b_key in cfg["buildings"].keys():
        b = resolve_building(cfg, b_key)
        for floor in b["floors"]:
            info_ws.cell(ir, 1, value=b_key)
            info_ws.cell(ir, 2, value=floor["name"])
            info_ws.cell(ir, 3, value=floor.get("plate") or "—")
            info_ws.cell(ir, 4, value=floor.get("wall") or "—")
            info_ws.cell(ir, 5, value=floor.get("ceiling") or "—")
            info_ws.cell(ir, 6, value=floor.get("note") or "")
            for c_idx in range(1, 7):
                info_ws.cell(ir, c_idx).border = thin_border()
                info_ws.cell(ir, c_idx).font = Font(size=10)
                info_ws.cell(ir, c_idx).alignment = Alignment(vertical="center", wrap_text=True)
            ir += 1
        if b.get("stairs"):
            for s in b["stairs"]:
                info_ws.cell(ir, 1, value=b_key)
                info_ws.cell(ir, 2, value="Лестницы")
                info_ws.cell(ir, 3, value=s)
                for c_idx in range(1, 7):
                    info_ws.cell(ir, c_idx).border = thin_border()
                    info_ws.cell(ir, c_idx).font = Font(size=10, italic=True, color="595959")
                ir += 1
        if b.get("ramps"):
            for r in b["ramps"]:
                info_ws.cell(ir, 1, value=b_key)
                info_ws.cell(ir, 2, value="Рампа")
                info_ws.cell(ir, 3, value=r)
                for c_idx in range(1, 7):
                    info_ws.cell(ir, c_idx).border = thin_border()
                    info_ws.cell(ir, c_idx).font = Font(size=10, italic=True, color="595959")
                ir += 1
    info_ws.column_dimensions["A"].width = 8
    info_ws.column_dimensions["B"].width = 14
    info_ws.column_dimensions["C"].width = 36
    info_ws.column_dimensions["D"].width = 36
    info_ws.column_dimensions["E"].width = 36
    info_ws.column_dimensions["F"].width = 40
    info_ws.freeze_panes = "A3"

    # ── сохранение ──
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Генератор пустого шаблона «Таблица 1-0 Репино.xlsx»")
    ap.add_argument("--config", type=Path, default=CONFIG_PATH, help="Путь к config.json")
    ap.add_argument("--out", type=Path, default=None,
                    help="Куда сохранить xlsx (по умолчанию — из config.output)")
    args = ap.parse_args()

    cfg = load_config(args.config)
    if args.out:
        out = args.out
    else:
        out_dir = Path(cfg["output"]["dir"])
        out = out_dir / cfg["output"]["filename"]

    build(cfg, out)
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    print(f"OK -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
