"""
build_1_0_extended_template.py — расширенный шаблон «Таблица 1-0 Репино».

Один лист, единая матрица:
- Слева 2 колонки: Корпус, Этаж/Отметка
- Зона А «До 0» (8 колонок) — заполняется только на строке «Цоколь» каждого корпуса
- Зона Б1 «Выше 0, конструктив» (5 колонок: Плита/Стены/Перекрытие/Лестницы/Рампа)
- Зона Б2 «Выше 0, отделка/инженерка» (9 колонок) — на каждом надземном этаже
- Зона В «По корпусу» (8 колонок) — заполняется один раз в строке «По корпусу в целом»
  каждого блока (Кровля/Окна/Витражи/Фасад/Лифты/МК/Двери/Сборн.каркас)
- Колонка «Итог» — % по применимым ячейкам строки
- Внизу — блок «Общие по объекту» (ИТП, Благоустройство, Нар.сети, Ввод)

Структура строк блока корпуса (К1..К6):
1. Заголовок «Корпус Кi»
2. «По корпусу в целом» — активна только зона В
3. Этажи: Цоколь, 1, 2, ..., Мансарда, Кровля
4. Лестницы (свои строки)
5. Рампы (только К2)
6. Итог по корпусу (формулы)

Использование:
    python build_1_0_extended_template.py
    python build_1_0_extended_template.py --out path\\to\\custom.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
CONFIG_PATH = ROOT / "config.json"
DEFAULT_OUT_NAME = "Таблица 1-0 расширенная Репино.xlsx"


# ── палитра ────────────────────────────────────────────────────────────────
C_TITLE_BG = "1F4E79"
C_TITLE_FG = "FFFFFF"
C_HEADER_FG = "FFFFFF"
C_HEADER_BG = "2E75B6"
C_GROUP_BG = "D9E1F2"
C_SUBGROUP_BG = "E7E6E6"
C_NA_BG = "D9D9D9"          # «не применимо»
C_NA_FG = "808080"
C_TOTAL_BG = "BDD7EE"
C_RED = "F8CBAD"
C_YELLOW = "FFE699"
C_GREEN = "C6EFCE"
C_CELL_DONE_BG = "E2EFD9"
C_CELL_DONE_FG = "00B050"
C_CELL_NOT_BG = "FBE4D5"
C_GRID = "BFBFBF"

# зоны (фон шапки 1-го уровня)
ZONE_BG = {
    "fix":         "595959",  # тёмно-серый  (закреплённые: Корпус/Этаж)
    "below_zero":  "ED7D31",  # оранжевый    (зона А «До 0»)
    "structural":  "1F4E79",  # тёмно-синий  (зона Б1 «Конструктив»)
    "finishing":   "2E75B6",  # синий        (зона Б2 «Отделка/инж.»)
    "building":    "548235",  # зелёный      (зона В «По корпусу»)
    "total":       "404040",  # графит       (Итог)
}

FONT_BASE = "Calibri"


# ── описания зон и колонок ─────────────────────────────────────────────────
# applies — где ячейка АКТИВНА (готова к 0/1). Везде остальное — серый N/A.
#   "basement"        — только строка «Цоколь»
#   "floor_plate"     — этаж, где config.floor.plate задан
#   "floor_wall"      — этаж, где config.floor.wall задан
#   "floor_ceiling"   — этаж, где config.floor.ceiling задан
#   "above_zero"      — этажи 1..N + Мансарда (без Цоколя, без Кровли, без Лест/Рамп)
#   "stair"           — строки лестниц
#   "ramp"            — строки рамп
#   "building_total"  — спец.строка «По корпусу в целом»
#   "k3_k6_only"      — только для корпусов К3..К6 (для «Сборн.каркас»)
#
# Колонка `key` уникальна по всему листу — пригодится позже на этапе заполнения.

ZONES = [
    {
        "id": "fix",
        "title": "",  # фиксированная зона без заголовка
        "columns": [
            {"key": "korpus", "title": "Корпус",      "width": 8,  "applies": "label"},
            {"key": "floor",  "title": "Этаж/Отметка","width": 32, "applies": "label"},
        ],
    },
    {
        "id": "below_zero",
        "title": "До 0",
        "columns": [
            {"key": "prep",       "title": "Подгот.",       "width": 7, "applies": "basement"},
            {"key": "pit",        "title": "Котлован",      "width": 7, "applies": "basement"},
            {"key": "sheetpile",  "title": "Шпунт",         "width": 7, "applies": "basement_k1_k3"},
            {"key": "crane",      "title": "БК",            "width": 6, "applies": "basement"},
            {"key": "wp_horiz",   "title": "ГИ\nгор.",      "width": 6, "applies": "basement"},
            {"key": "wp_vert",    "title": "ГИ\nверт.",     "width": 6, "applies": "basement"},
            {"key": "backfill",   "title": "Обр.\nзасыпка", "width": 7, "applies": "basement"},
            {"key": "ins_below",  "title": "Утепл.\nниже 0","width": 7, "applies": "basement"},
        ],
    },
    {
        "id": "structural",
        "title": "Выше 0 — конструктив",
        "columns": [
            {"key": "plate",     "title": "Плита",      "width": 8, "applies": "floor_plate"},
            {"key": "wall",      "title": "Стены",      "width": 8, "applies": "floor_wall"},
            {"key": "ceiling",   "title": "Перекр.",    "width": 8, "applies": "floor_ceiling"},
            {"key": "stair",     "title": "Лестн.",     "width": 8, "applies": "stair"},
            {"key": "ramp",      "title": "Рампа",      "width": 8, "applies": "ramp"},
        ],
    },
    {
        "id": "building",
        "title": "По корпусу",
        "columns": [
            {"key": "roof",      "title": "Кровля",      "width": 8, "applies": "building_total"},
            {"key": "windows",   "title": "Окна Al",     "width": 8, "applies": "building_total_no_k2"},
            {"key": "stained",   "title": "Витражи",     "width": 8, "applies": "building_total"},
            {"key": "facade",    "title": "Фасад",       "width": 8, "applies": "building_total"},
            {"key": "lifts",     "title": "Лифты",       "width": 7, "applies": "building_total"},
            {"key": "metal",     "title": "МК",          "width": 6, "applies": "building_total"},
            {"key": "doors",     "title": "Двери",       "width": 7, "applies": "building_total"},
            {"key": "frame",     "title": "Сборн.\nкаркас","width": 8, "applies": "k3_k6_only"},
        ],
    },
    {
        "id": "finishing",
        "title": "Выше 0 — отделка / инженерка",
        "columns": [
            {"key": "partitions","title": "Перегор.",   "width": 8, "applies": "above_zero"},
            {"key": "screed",    "title": "Стяжка",     "width": 7, "applies": "above_zero"},
            {"key": "heating",   "title": "Отопл.",     "width": 7, "applies": "above_zero"},
            {"key": "water",     "title": "Водосн.",    "width": 7, "applies": "above_zero"},
            {"key": "sewage",    "title": "Водоотв.",   "width": 7, "applies": "above_zero"},
            {"key": "vent",      "title": "Вентил.",    "width": 7, "applies": "above_zero"},
            {"key": "elec",      "title": "ЭС",         "width": 6, "applies": "above_zero"},
            {"key": "weak",      "title": "Слаб.\nточ.","width": 7, "applies": "above_zero"},
            {"key": "finish",    "title": "Отделка",    "width": 8, "applies": "above_zero"},
        ],
    },
    {
        "id": "total",
        "title": "",
        "columns": [
            {"key": "total", "title": "Итог", "width": 9, "applies": "formula"},
        ],
    },
]

# Виды работ объектного уровня (отдельный блок снизу)
OBJECT_WORKS = [
    {"key": "itp",         "title": "ИТП"},
    {"key": "landscaping", "title": "Благоустройство"},
    {"key": "ext_nets",    "title": "Наружные инженерные сети"},
    {"key": "handover",    "title": "Ввод объекта в эксплуатацию"},
]


# ── утилиты ────────────────────────────────────────────────────────────────
def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def thin_border() -> Border:
    s = Side(style="thin", color=C_GRID)
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border() -> Border:
    s = Side(style="medium", color="404040")
    return Border(left=s, right=s, top=s, bottom=s)


def load_config(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def resolve_building(cfg: dict, key: str) -> dict:
    b = cfg["buildings"][key]
    if "extends" in b:
        tpl_key = b["extends"]
        tpl = cfg["templates"][tpl_key]
        merged = dict(tpl)
        for k, v in b.items():
            if k != "extends":
                merged[k] = v
        return merged
    return b


# ── строки блоков корпусов ────────────────────────────────────────────────
def cell_applies(applies: str, row_kind: str, b_key: str, floor: dict | None) -> bool:
    """row_kind: building_total | floor | stair | ramp.
    Возвращает True, если ячейка АКТИВНА (готова к 0/1) на этой строке."""
    if applies == "label" or applies == "formula":
        return True
    if applies == "basement":
        return row_kind == "floor" and floor is not None and floor["name"] == "Цоколь"
    if applies == "basement_k1_k3":
        return (row_kind == "floor" and floor is not None and floor["name"] == "Цоколь"
                and b_key in ("К1", "К2", "К3"))
    if applies == "floor_plate":
        return row_kind == "floor" and bool(floor and floor.get("plate"))
    if applies == "floor_wall":
        return row_kind == "floor" and bool(floor and floor.get("wall"))
    if applies == "floor_ceiling":
        return row_kind == "floor" and bool(floor and floor.get("ceiling"))
    if applies == "above_zero":
        if row_kind != "floor" or floor is None:
            return False
        name = floor["name"]
        return name not in ("Цоколь", "Кровля")
    if applies == "stair":
        return row_kind == "stair"
    if applies == "ramp":
        return row_kind == "ramp"
    if applies == "building_total":
        return row_kind == "building_total"
    if applies == "building_total_no_k2":
        return row_kind == "building_total" and b_key != "К2"
    if applies == "k3_k6_only":
        return row_kind == "building_total" and b_key in ("К3", "К4", "К5", "К6")
    return False


# ── основной билдер ────────────────────────────────────────────────────────
def build(cfg: dict, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица 1-0"

    # 1) Раскладка колонок
    col_map: list[dict] = []          # плоский список колонок с зоной и индексом
    zone_spans: list[tuple[str, int, int]] = []  # (zone_id, first_col, last_col)
    col_idx = 1
    for zone in ZONES:
        z_first = col_idx
        for c in zone["columns"]:
            col_map.append({"zone": zone["id"], "col_idx": col_idx, **c})
            col_idx += 1
        zone_spans.append((zone["id"], z_first, col_idx - 1))
    n_cols = col_idx - 1

    # Колонки по семантическим зонам — пригодится для формул итогов
    def cols_in_zone(zone_id: str) -> list[int]:
        return [c["col_idx"] for c in col_map if c["zone"] == zone_id]

    DATA_ZONES = ("below_zero", "structural", "finishing", "building")
    DATA_FIRST = min(min(cols_in_zone(z)) for z in DATA_ZONES)
    DATA_LAST  = max(max(cols_in_zone(z)) for z in DATA_ZONES)
    COL_TOTAL  = cols_in_zone("total")[0]

    # 2) Шапка
    # строка 1 — заголовок таблицы
    title_text = f"Таблица 1-0 — {cfg['project_full_name']} — расширенная (по всем видам работ)"
    ws.cell(1, 1, value=title_text).font = Font(name=FONT_BASE, size=14, bold=True, color=C_TITLE_FG)
    ws.cell(1, 1).fill = fill(C_TITLE_BG)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26

    # строка 2 — легенда
    legend = "Легенда:   1 = готово   ·   0 = не готово   ·   пусто = не требуется"
    ws.cell(2, 1, value=legend).font = Font(name=FONT_BASE, size=9, italic=True, color="595959")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 16

    # строки 3-4 — шапка двух уровней
    HEADER_ROW_1 = 3
    HEADER_ROW_2 = 4

    for z_id, c_first, c_last in zone_spans:
        zone = next(z for z in ZONES if z["id"] == z_id)
        title = zone["title"]
        bg = ZONE_BG[z_id]
        if not title:
            # для fix/total — мерджим вертикально (на обе строки) и пишем titles нижней строки в HEADER_ROW_1
            for c_idx in range(c_first, c_last + 1):
                col_def = next(c for c in col_map if c["col_idx"] == c_idx)
                ws.cell(HEADER_ROW_1, c_idx, value=col_def["title"])
                ws.merge_cells(start_row=HEADER_ROW_1, start_column=c_idx,
                               end_row=HEADER_ROW_2, end_column=c_idx)
                cell = ws.cell(HEADER_ROW_1, c_idx)
                cell.fill = fill(bg)
                cell.font = Font(name=FONT_BASE, size=10, bold=True, color=C_HEADER_FG)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border()
        else:
            # верх — заголовок зоны (merged)
            ws.cell(HEADER_ROW_1, c_first, value=title)
            if c_first != c_last:
                ws.merge_cells(start_row=HEADER_ROW_1, start_column=c_first,
                               end_row=HEADER_ROW_1, end_column=c_last)
            cell = ws.cell(HEADER_ROW_1, c_first)
            cell.fill = fill(bg)
            cell.font = Font(name=FONT_BASE, size=11, bold=True, color=C_HEADER_FG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

            # низ — имена колонок
            for c_idx in range(c_first, c_last + 1):
                col_def = next(c for c in col_map if c["col_idx"] == c_idx)
                cell = ws.cell(HEADER_ROW_2, c_idx, value=col_def["title"])
                cell.fill = fill(C_HEADER_BG)
                cell.font = Font(name=FONT_BASE, size=9, bold=True, color=C_HEADER_FG)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border()

    ws.row_dimensions[HEADER_ROW_1].height = 22
    ws.row_dimensions[HEADER_ROW_2].height = 34

    # ширины
    for c in col_map:
        ws.column_dimensions[get_column_letter(c["col_idx"])].width = c["width"]

    # 3) Блоки корпусов
    row = HEADER_ROW_2 + 1
    building_total_rows: list[tuple[str, int]] = []
    building_data_ranges: list[tuple[str, int, int]] = []  # для итогов по объекту

    KORPUS_COL = next(c["col_idx"] for c in col_map if c["key"] == "korpus")
    FLOOR_COL  = next(c["col_idx"] for c in col_map if c["key"] == "floor")

    def write_label(r: int, korpus: str, floor_label: str, kind: str) -> None:
        ws.cell(r, KORPUS_COL, value=korpus)
        ws.cell(r, FLOOR_COL,  value=floor_label)
        for c_idx in (KORPUS_COL, FLOOR_COL):
            cell = ws.cell(r, c_idx)
            cell.font = Font(name=FONT_BASE, size=10,
                             bold=(kind == "building_total"))
            cell.alignment = Alignment(
                horizontal="center" if c_idx == KORPUS_COL else "left",
                vertical="center", wrap_text=True, indent=0 if c_idx == KORPUS_COL else 1,
            )
            cell.border = thin_border()
            if kind == "building_total":
                cell.fill = fill(C_TOTAL_BG)

    def write_data_cell(r: int, col_def: dict, kind: str, b_key: str, floor: dict | None) -> None:
        applies = col_def["applies"]
        active = cell_applies(applies, kind, b_key, floor)
        cell = ws.cell(r, col_def["col_idx"])
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BASE, size=10)
        if not active:
            cell.fill = fill(C_NA_BG)

    for b_key in cfg["buildings"].keys():
        b = resolve_building(cfg, b_key)

        # 3.1) Заголовок корпуса (одна merged-строка)
        title_row = row
        ws.cell(row, 1, value=f"Корпус {b_key}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(row, 1)
        cell.fill = fill(C_GROUP_BG)
        cell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = thin_border()
        ws.row_dimensions[row].height = 22
        row += 1

        # 3.2) Строка «По корпусу в целом»
        bt_row = row
        write_label(bt_row, b_key, "По корпусу в целом", "building_total")
        for c in col_map:
            if c["zone"] not in DATA_ZONES:
                continue
            write_data_cell(bt_row, c, "building_total", b_key, None)
        row += 1

        # 3.3) Этажи
        floor_data_rows: list[int] = []
        for floor in b["floors"]:
            r = row
            floor_data_rows.append(r)
            label = floor["name"]
            if floor.get("elev"):
                label += f"   (отм. {floor['elev']})"
            write_label(r, b_key, label, "floor")
            for c in col_map:
                if c["zone"] not in DATA_ZONES:
                    continue
                write_data_cell(r, c, "floor", b_key, floor)
            row += 1

        # 3.4) Лестницы
        stair_rows: list[int] = []
        if b.get("stairs"):
            for s_name in b["stairs"]:
                r = row
                stair_rows.append(r)
                write_label(r, b_key, f"Лестница: {s_name}", "stair")
                for c in col_map:
                    if c["zone"] not in DATA_ZONES:
                        continue
                    write_data_cell(r, c, "stair", b_key, None)
                row += 1

        # 3.5) Рампы
        ramp_rows: list[int] = []
        if b.get("ramps"):
            for r_name in b["ramps"]:
                r = row
                ramp_rows.append(r)
                write_label(r, b_key, f"Рампа: {r_name}", "ramp")
                for c in col_map:
                    if c["zone"] not in DATA_ZONES:
                        continue
                    write_data_cell(r, c, "ramp", b_key, None)
                row += 1

        # 3.6) Итог-строка по корпусу — формулы
        total_row = row
        building_total_rows.append((b_key, total_row))
        all_data_rows = [bt_row] + floor_data_rows + stair_rows + ramp_rows
        building_data_ranges.append((b_key, min(all_data_rows), max(all_data_rows)))

        write_label(total_row, b_key, f"Итог по корпусу {b_key}", "building_total")
        ws.cell(total_row, FLOOR_COL).font = Font(name=FONT_BASE, size=11, bold=True, color="1F4E79")

        r_first, r_last = min(all_data_rows), max(all_data_rows)
        for c in col_map:
            if c["zone"] not in DATA_ZONES:
                continue
            col_letter = get_column_letter(c["col_idx"])
            formula = (
                f'=IF(COUNT({col_letter}{r_first}:{col_letter}{r_last})=0,"",'
                f'SUM({col_letter}{r_first}:{col_letter}{r_last})/'
                f'COUNT({col_letter}{r_first}:{col_letter}{r_last}))'
            )
            cell = ws.cell(total_row, c["col_idx"], value=formula)
            cell.number_format = "0%"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name=FONT_BASE, size=10, bold=True)
            cell.fill = fill(C_TOTAL_BG)
            cell.border = medium_border()

        # Итог-строки этой корпусной строки (Итог колонки)
        rng = (
            f"{get_column_letter(DATA_FIRST)}{r_first}:"
            f"{get_column_letter(DATA_LAST)}{r_last}"
        )
        # При суммировании по диапазону формулы дают «—»-ячейки как ошибку.
        # Используем SUMPRODUCT/ISNUMBER для подсчёта только числовых.
        grand_formula = (
            f'=IF(SUMPRODUCT(--ISNUMBER({rng}))=0,"",'
            f'SUMPRODUCT(--ISNUMBER({rng})*{rng})/SUMPRODUCT(--ISNUMBER({rng})))'
        )
        cell = ws.cell(total_row, COL_TOTAL, value=grand_formula)
        cell.number_format = "0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
        cell.fill = fill(C_TOTAL_BG)
        cell.border = medium_border()
        ws.row_dimensions[total_row].height = 22
        row += 1

        # пустая разделительная
        ws.row_dimensions[row].height = 6
        row += 1

    # 4) Блок «Общие по объекту»
    row += 1
    obj_title_row = row
    ws.cell(row, 1, value="Общие по объекту (виды работ объектного уровня)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row, 1)
    cell.fill = fill(C_GROUP_BG)
    cell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    object_work_rows: list[int] = []
    for work in OBJECT_WORKS:
        object_work_rows.append(row)
        write_label(row, "", work["title"], "stair")
        ws.cell(row, FLOOR_COL).font = Font(name=FONT_BASE, size=10, bold=True)
        # data-колонки — серые: эти виды не разносятся ни по этажам, ни по корпусам
        for c in col_map:
            if c["zone"] not in DATA_ZONES:
                continue
            cell = ws.cell(row, c["col_idx"])
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(C_NA_BG)
        # колонка «Итог» — активна для ручного ввода 1/0
        cell = ws.cell(row, COL_TOTAL)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BASE, size=10)
        # без значения — пользователь вводит 1/0; CF подкрасит
        row += 1

    # 4.1) «ИТОГО общестроительные» — среднее по 4 общим
    objtotal_row = row
    write_label(objtotal_row, "", "ИТОГО общестроительные", "building_total")
    ws.cell(objtotal_row, FLOOR_COL).font = Font(name=FONT_BASE, size=11, bold=True, color="1F4E79")
    # data-колонки серые
    for c in col_map:
        if c["zone"] not in DATA_ZONES:
            continue
        cell = ws.cell(objtotal_row, c["col_idx"])
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill(C_NA_BG)
    obj_refs = [f"{get_column_letter(COL_TOTAL)}{r}" for r in object_work_rows]
    cell = ws.cell(objtotal_row, COL_TOTAL,
                   value=f'=IFERROR(AVERAGE({",".join(obj_refs)}),"")')
    cell.number_format = "0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BASE, size=12, bold=True, color="1F4E79")
    cell.fill = fill(C_TOTAL_BG)
    cell.border = medium_border()
    ws.row_dimensions[objtotal_row].height = 22
    row += 1

    # пустая разделительная
    ws.row_dimensions[row].height = 6
    row += 1

    # 5) «ИТОГО по объекту Репино» — учитывает 6 корпусов + общестроительные
    grand_row = row
    write_label(grand_row, "", "ИТОГО по объекту Репино", "building_total")
    ws.cell(grand_row, FLOOR_COL).font = Font(name=FONT_BASE, size=12, bold=True, color="FFFFFF")
    ws.cell(grand_row, FLOOR_COL).fill = fill(C_TITLE_BG)
    ws.cell(grand_row, KORPUS_COL).fill = fill(C_TITLE_BG)

    # в data-колонках — среднее по итогам корпусов (общестроительных в data-колонках нет)
    for c in col_map:
        if c["zone"] not in DATA_ZONES:
            continue
        col_letter = get_column_letter(c["col_idx"])
        refs = [f"{col_letter}{r}" for _, r in building_total_rows]
        formula = f'=IFERROR(AVERAGE({",".join(refs)}),"")'
        cell = ws.cell(grand_row, c["col_idx"], value=formula)
        cell.number_format = "0%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_BASE, size=11, bold=True, color="FFFFFF")
        cell.fill = fill(C_TITLE_BG)
        cell.border = medium_border()

    # в Итог — среднее по 6 корпусам + 1 общестроит
    all_total_refs = [f"{get_column_letter(COL_TOTAL)}{r}" for _, r in building_total_rows]
    all_total_refs.append(f"{get_column_letter(COL_TOTAL)}{objtotal_row}")
    cell = ws.cell(grand_row, COL_TOTAL,
                   value=f'=IFERROR(AVERAGE({",".join(all_total_refs)}),"")')
    cell.number_format = "0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name=FONT_BASE, size=14, bold=True, color="FFFFFF")
    cell.fill = fill(C_TITLE_BG)
    cell.border = medium_border()
    ws.row_dimensions[grand_row].height = 28
    row += 1

    # 6) Условное форматирование
    # 6.1) Дискретное окрашивание data-ячеек 1/0
    data_cf_range_first_row = HEADER_ROW_2 + 1
    data_cf_range_last_row = grand_row
    data_cf_range = (
        f"{get_column_letter(DATA_FIRST)}{data_cf_range_first_row}:"
        f"{get_column_letter(DATA_LAST)}{data_cf_range_last_row}"
    )
    dxf_done = DifferentialStyle(
        font=Font(color=C_CELL_DONE_FG, bold=True),
        fill=fill(C_CELL_DONE_BG),
    )
    dxf_not = DifferentialStyle(fill=fill(C_CELL_NOT_BG))
    ws.conditional_formatting.add(
        data_cf_range,
        Rule(type="cellIs", operator="equal", formula=["1"], dxf=dxf_done, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        data_cf_range,
        Rule(type="cellIs", operator="equal", formula=["0"], dxf=dxf_not, stopIfTrue=False),
    )

    # 6.2) Цветная шкала на итогах (по колонке «Итог», по итогам корпусов, общий)
    th = cfg.get("thresholds_pct", {"red_below": 50, "yellow_below": 100})
    red_below = th["red_below"] / 100.0
    yellow_below = th["yellow_below"] / 100.0

    def add_pct_rules(rng: str) -> None:
        for op, formula, color in (
            ("lessThan", str(red_below), C_RED),
            ("lessThan", str(yellow_below), C_YELLOW),
            ("greaterThanOrEqual", str(yellow_below), C_GREEN),
        ):
            ws.conditional_formatting.add(
                rng, CellIsRule(operator=op, formula=[formula], fill=fill(color))
            )

    # колонка Итог по строкам (включая Общие по объекту и оба ИТОГО)
    add_pct_rules(f"{get_column_letter(COL_TOTAL)}{HEADER_ROW_2+1}:"
                  f"{get_column_letter(COL_TOTAL)}{grand_row}")
    # итоги корпусов и общий — по data-колонкам
    for _, tr in building_total_rows:
        add_pct_rules(f"{get_column_letter(DATA_FIRST)}{tr}:"
                      f"{get_column_letter(DATA_LAST)}{tr}")
    add_pct_rules(f"{get_column_letter(DATA_FIRST)}{grand_row}:"
                  f"{get_column_letter(DATA_LAST)}{grand_row}")

    # 7) Заморозка шапки + лево
    ws.freeze_panes = f"{get_column_letter(DATA_FIRST)}{HEADER_ROW_2 + 1}"

    # 8) Печать
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"1:{HEADER_ROW_2}"

    # 9) Сохранение
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Расширенный шаблон «Таблица 1-0 Репино»")
    ap.add_argument("--config", type=Path, default=CONFIG_PATH)
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()

    cfg = load_config(args.config)
    if args.out:
        out = args.out
    else:
        out = Path(cfg["output"]["dir"]) / DEFAULT_OUT_NAME

    build(cfg, out)
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    print(f"OK -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
