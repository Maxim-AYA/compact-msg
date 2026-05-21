"""Сборка xlsx «Численность подрядчиков на площадке - Репино» (ежедневный табель v3).

Структура листа «Численность по дням»:
- Колонки: даты (12 месяцев × 28-31 день), после каждого месяца — «ИТОГО мес.».
- Строки: 8 подрядчиков × 3 (ИТР / Раб / Σ) + ИТОГО × 3 (ИТР / Раб / Σ).
- Группировка по столбцам: дни сворачиваются в outline, остаётся только колонка ИТОГО месяца.
- Выходные подсвечены, шапка темно-синяя.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from datetime import date
import calendar, shutil, os, tempfile

OUT_FINAL = r"C:\Авраменко\Claude Code Projects\МСГ\Подрядчики\Численность подрядчиков на площадке - Репино.xlsx"

# === Подрядчики ===
contractors = [
    ("АО СК \"Компакт\"",    "Генподрядчик"),
    ("АО \"Компакт\"",       "Монолит к. 5-6"),
    ("ООО \"Болверк-Норд\"", "Шпунт"),
    ("ООО \"Авангард\"",     "Разработка котлована"),
    ("ООО \"Дельта-Строй\"", "Гидроизоляция"),
    ("ООО \"Ладога\"",       "Монолит к. 3-4"),
    ("ООО \"ШефСтрой\"",     "Монолит к. 1-2"),
    ("ООО \"Покровъ\"",      "Перегородки"),
]
contr_short = ['АО СК "Компакт"', 'АО "Компакт"', 'ООО "Болверк-Норд"',
               'ООО "Авангард"', 'ООО "Дельта-Строй"', 'ООО "Ладога"',
               'ООО "ШефСтрой"', 'ООО "Покровъ"']
# Виды работ — точно как в столбце СМР файла «Контакты Генподрядчика и подрядчиков».
# Для АО «Компакт» (нет в контактах) — формулировка по аналогии с Ладога/ШефСтрой.
contr_zone  = ["Генподрядчик",
               "Подрядчик по монолиту корп 5,6",
               "Подрядчик по шпунту",
               "Подрядчик по разработке котлована",
               "Подрядчик по гидроизоляции",
               "Подрядчик по монолиту корп 3,4",
               "Подрядчик по монолиту корп 1,2",
               "Подрядчик по перегородкам"]
N_CONTR = len(contractors)

# === Стили ===
NAVY = "1F4E79"
LIGHT_BLUE = "DCE6F1"
LIGHTER_BLUE = "EAF1F8"
GREY_LIGHT = "F2F2F2"
WEEKEND_FILL = "FFF2CC"  # светло-бежевый для выходных

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
DATA_BOLD = Font(name="Calibri", size=10, bold=True)
THIN = Side(border_style="thin", color="A6A6A6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NW = Alignment(horizontal="center", vertical="center")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
SUM_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SUB_FILL = PatternFill("solid", fgColor=LIGHTER_BLUE)
GREY_FILL = PatternFill("solid", fgColor=GREY_LIGHT)
WEEKEND = PatternFill("solid", fgColor=WEEKEND_FILL)

wb = Workbook()

# ============================================================
# Лист «Справочник»
# ============================================================
ws_ref = wb.active
ws_ref.title = "Справочник"
ws_ref.cell(row=1, column=1, value="№")
ws_ref.cell(row=1, column=2, value="Организация")
ws_ref.cell(row=1, column=3, value="Зона ответственности (СМР)")
for col in range(1, 4):
    c = ws_ref.cell(row=1, column=col)
    c.fill = NAVY_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

for i, (org, smr) in enumerate(contractors):
    r = i + 2
    ws_ref.cell(row=r, column=1, value=i + 1).alignment = CENTER_NW
    ws_ref.cell(row=r, column=2, value=org)
    ws_ref.cell(row=r, column=3, value=smr)
    for col in range(1, 4):
        c = ws_ref.cell(row=r, column=col)
        c.font = DATA_FONT
        c.border = BORDER
        if col > 1:
            c.alignment = Alignment(vertical="center")

ws_ref.column_dimensions["A"].width = 5
ws_ref.column_dimensions["B"].width = 26
ws_ref.column_dimensions["C"].width = 32
ws_ref.row_dimensions[1].height = 28
for r in range(2, 2 + N_CONTR):
    ws_ref.row_dimensions[r].height = 20

# ============================================================
# Лист «Численность по дням»
# ============================================================
ws = wb.create_sheet("Численность по дням")

# Период: 12 месяцев с мая 2026
months = []
y, m = 2026, 5
for _ in range(12):
    months.append((y, m))
    m += 1
    if m > 12:
        m = 1; y += 1

months_ru = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг",
             "Сен", "Окт", "Ноя", "Дек"]
wd_ru = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]

HDR_MONTH = 1
HDR_DAY = 2
HDR_WD = 3
DATA_R0 = 4

# Шапка слева (A, B): «Подрядчик» / «Кат.»
ws.cell(row=HDR_MONTH, column=1, value="Подрядчик")
ws.merge_cells(start_row=HDR_MONTH, start_column=1, end_row=HDR_WD, end_column=1)
ws.cell(row=HDR_MONTH, column=2, value="Кат.")
ws.merge_cells(start_row=HDR_MONTH, start_column=2, end_row=HDR_WD, end_column=2)
for col in (1, 2):
    c = ws.cell(row=HDR_MONTH, column=col)
    c.fill = NAVY_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

# Построение колонок: дни месяца → ИТОГО месяца → дни следующего месяца → ...
month_cols = {}  # (y,m) -> dict(days=[(col,d,wd),...], total_col=int)
col = 3
for (yy, mm) in months:
    ndays = calendar.monthrange(yy, mm)[1]
    days = []
    for d in range(1, ndays + 1):
        wd = date(yy, mm, d).weekday()
        days.append((col, d, wd))
        col += 1
    total_col = col
    col += 1
    month_cols[(yy, mm)] = {"days": days, "total_col": total_col, "ndays": ndays}

last_col = col - 1

# Шапка по дням и месяцам
for (yy, mm), info in month_cols.items():
    start_col = info["days"][0][0]
    end_col = info["total_col"]

    # Row 1: название месяца (merged по дням + итог)
    ws.cell(row=HDR_MONTH, column=start_col, value=f"{months_ru[mm-1]} {yy}")
    ws.merge_cells(start_row=HDR_MONTH, start_column=start_col,
                   end_row=HDR_MONTH, end_column=end_col)
    mc = ws.cell(row=HDR_MONTH, column=start_col)
    mc.font = HEADER_FONT
    mc.alignment = CENTER
    mc.fill = NAVY_FILL
    mc.border = BORDER

    # Row 2: числа + ИТОГО
    for c, d, wd in info["days"]:
        cell = ws.cell(row=HDR_DAY, column=c, value=d)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_NW
        cell.fill = NAVY_FILL
        cell.border = BORDER
    cell = ws.cell(row=HDR_DAY, column=info["total_col"], value="Средн.")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_NW
    cell.fill = NAVY_FILL
    cell.border = BORDER

    # Row 3: день недели + «мес.»
    for c, d, wd in info["days"]:
        cell = ws.cell(row=HDR_WD, column=c, value=wd_ru[wd])
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_NW
        cell.fill = NAVY_FILL
        cell.border = BORDER
    cell = ws.cell(row=HDR_WD, column=info["total_col"], value="числ.")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_NW
    cell.fill = NAVY_FILL
    cell.border = BORDER

# Строки подрядчиков (8 × 3)
contr_rows = {}
r = DATA_R0
bold_inline = InlineFont(rFont="Calibri", sz=10, b=True, color="000000")
reg_inline  = InlineFont(rFont="Calibri", sz=10, b=False, color="000000")
TECH_FILL = PatternFill("solid", fgColor="E2EFDA")  # светло-зелёный для строки «Техника»
for i in range(N_CONTR):
    name = contr_short[i]
    zone = contr_zone[i]
    rich = CellRichText([
        TextBlock(bold_inline, name),
        TextBlock(reg_inline, f"\n({zone})"),
    ])
    ws.cell(row=r, column=1, value=rich)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=1)
    nc = ws.cell(row=r, column=1)
    nc.alignment = Alignment(vertical="center", horizontal="left",
                             wrap_text=True, indent=1)
    nc.fill = GREY_FILL
    nc.border = BORDER

    ws.cell(row=r,     column=2, value="ИТР")
    ws.cell(row=r + 1, column=2, value="Раб")
    ws.cell(row=r + 2, column=2, value="Σ")
    ws.cell(row=r + 3, column=2, value="Техн.")
    for offset, fill in ((0, SUB_FILL), (1, SUB_FILL), (2, SUM_FILL), (3, TECH_FILL)):
        cc = ws.cell(row=r + offset, column=2)
        cc.font = DATA_BOLD if offset >= 2 else DATA_FONT
        cc.alignment = CENTER_NW
        cc.border = BORDER
        cc.fill = fill

    contr_rows[name] = {"itr": r, "rab": r + 1, "sum": r + 2, "tech": r + 3}
    r += 4

# ИТОГО снизу
tot_itr  = r
tot_rab  = r + 1
tot_sum  = r + 2
tot_tech = r + 3
ws.cell(row=tot_itr, column=1, value="ИТОГО")
ws.merge_cells(start_row=tot_itr, start_column=1, end_row=tot_tech, end_column=1)
tc = ws.cell(row=tot_itr, column=1)
tc.font = HEADER_FONT
tc.alignment = CENTER
tc.fill = NAVY_FILL
tc.border = BORDER

ws.cell(row=tot_itr,  column=2, value="ИТР")
ws.cell(row=tot_rab,  column=2, value="Раб")
ws.cell(row=tot_sum,  column=2, value="Σ")
ws.cell(row=tot_tech, column=2, value="Техн.")
for rr, fill in ((tot_itr, SUM_FILL), (tot_rab, SUM_FILL),
                 (tot_sum, SUM_FILL), (tot_tech, TECH_FILL)):
    cc = ws.cell(row=rr, column=2)
    cc.font = DATA_BOLD
    cc.alignment = CENTER_NW
    cc.fill = fill
    cc.border = BORDER

# === Данные ===
# 18.05.2026 — стартовая запись
may_data = {
    'АО СК "Компакт"':    (8, None),
    'АО "Компакт"':       (None, 27),
    'ООО "Болверк-Норд"': (None, None),
    'ООО "Авангард"':     (2, 2),
    'ООО "Дельта-Строй"': (1, 4),
    'ООО "Ладога"':       (2, 32),
    'ООО "ШефСтрой"':     (1, 6),
    'ООО "Покровъ"':      (1, 1),
}

for name in contr_short:
    rows = contr_rows[name]
    itr_r, rab_r, sum_r, tech_r = rows["itr"], rows["rab"], rows["sum"], rows["tech"]

    for (yy, mm), info in month_cols.items():
        for c, d, wd in info["days"]:
            L = get_column_letter(c)
            if yy == 2026 and mm == 5 and d == 18:
                itr, rab = may_data[name]
                if itr is not None:
                    ws.cell(row=itr_r, column=c, value=itr)
                if rab is not None:
                    ws.cell(row=rab_r, column=c, value=rab)
            ws.cell(row=sum_r, column=c,
                    value=f'=IF(AND({L}{itr_r}="",{L}{rab_r}=""),"",N({L}{itr_r})+N({L}{rab_r}))')
            # ячейка «Техника» — пустая, заполняется вручную

        # «Средн.» месяца по строке подрядчика — среднее по дням с >0 (простои не учитываем)
        tc_col = info["total_col"]
        first_L = get_column_letter(info["days"][0][0])
        last_L  = get_column_letter(info["days"][-1][0])
        def avg_formula(row):
            return f'=IFERROR(ROUND(AVERAGEIF({first_L}{row}:{last_L}{row},">0"),1),"")'
        ws.cell(row=itr_r,  column=tc_col, value=avg_formula(itr_r))
        ws.cell(row=rab_r,  column=tc_col, value=avg_formula(rab_r))
        ws.cell(row=sum_r,  column=tc_col, value=avg_formula(sum_r))
        ws.cell(row=tech_r, column=tc_col, value=avg_formula(tech_r))

# ИТОГО строки (по всем подрядчикам)
all_itr_rows  = [contr_rows[n]["itr"]  for n in contr_short]
all_rab_rows  = [contr_rows[n]["rab"]  for n in contr_short]
all_tech_rows = [contr_rows[n]["tech"] for n in contr_short]

for (yy, mm), info in month_cols.items():
    # Дневные колонки — сумма по 8 подрядчикам.
    for c, _, _ in info["days"]:
        L = get_column_letter(c)
        itr_cells  = ",".join(f"{L}{rr}" for rr in all_itr_rows)
        rab_cells  = ",".join(f"{L}{rr}" for rr in all_rab_rows)
        tech_cells = ",".join(f"{L}{rr}" for rr in all_tech_rows)
        ws.cell(row=tot_itr,  column=c, value=f"=SUM({itr_cells})")
        ws.cell(row=tot_rab,  column=c, value=f"=SUM({rab_cells})")
        ws.cell(row=tot_sum,  column=c, value=f'=IF(AND({L}{tot_itr}="",{L}{tot_rab}=""),"",N({L}{tot_itr})+N({L}{tot_rab}))')
        ws.cell(row=tot_tech, column=c, value=f"=SUM({tech_cells})")
    # Колонка «Средн.» месяца — среднее по непустым/положительным дням.
    tc_col = info["total_col"]
    first_L = get_column_letter(info["days"][0][0])
    last_L  = get_column_letter(info["days"][-1][0])
    for rr in (tot_itr, tot_rab, tot_sum, tot_tech):
        ws.cell(row=rr, column=tc_col,
                value=f'=IFERROR(ROUND(AVERAGEIF({first_L}{rr}:{last_L}{rr},">0"),1),"")')

# === Стили ячеек данных + outline ===
# В блоке подрядчика 4 строки: 0=ИТР, 1=Раб, 2=Σ, 3=Техника.
# Принцип подсветки: формулы и метки — цветные; ячейки для ручного ввода — белые.
sum_rows_set = {contr_rows[n]["sum"]  for n in contr_short} | {tot_sum}
tech_contr_rows = {contr_rows[n]["tech"] for n in contr_short}  # только подрядчики (ввод)

for (yy, mm), info in month_cols.items():
    for c, d, wd in info["days"]:
        L = get_column_letter(c)
        ws.column_dimensions[L].outline_level = 1
        ws.column_dimensions[L].width = 4.2
        is_weekend = wd >= 5
        for rr in range(DATA_R0, tot_tech + 1):
            cell = ws.cell(row=rr, column=c)
            cell.font = DATA_BOLD if rr in sum_rows_set or rr >= tot_itr else DATA_FONT
            cell.alignment = CENTER_NW
            cell.border = BORDER
            if is_weekend:
                cell.fill = WEEKEND
            elif rr == tot_tech:
                cell.fill = TECH_FILL          # ИТОГО техники по дню — формула, оставляем зелёный
            elif rr in tech_contr_rows:
                pass                            # техника подрядчика — ввод, без заливки
            elif rr in sum_rows_set:
                cell.fill = SUB_FILL            # Σ — формула
            elif rr == tot_itr or rr == tot_rab:
                cell.fill = SUB_FILL            # ИТОГО ИТР/Раб — формулы
        # подсветка выходных в шапке дня недели
        if is_weekend:
            wdc = ws.cell(row=HDR_WD, column=c)
            wdc.fill = PatternFill("solid", fgColor="B85450")

    # ИТОГО колонка месяца — пошире
    tc_col = info["total_col"]
    ws.column_dimensions[get_column_letter(tc_col)].width = 8
    for rr in range(DATA_R0, tot_tech + 1):
        cell = ws.cell(row=rr, column=tc_col)
        cell.fill = TECH_FILL if (rr in tech_contr_rows or rr == tot_tech) else SUM_FILL
        cell.font = DATA_BOLD
        cell.alignment = CENTER_NW
        cell.border = BORDER

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 5

# Сворачиваем все месяцы кроме мая 2026 — скрываем дневные колонки,
# а колонку «ИТОГО мес.» помечаем collapsed=True (маркер «+» справа).
for (yy, mm), info in month_cols.items():
    if yy == 2026 and mm == 5:
        continue
    for c, d, wd in info["days"]:
        ws.column_dimensions[get_column_letter(c)].hidden = True
    ws.column_dimensions[get_column_letter(info["total_col"])].collapsed = True

ws.row_dimensions[HDR_MONTH].height = 22
ws.row_dimensions[HDR_DAY].height = 18
ws.row_dimensions[HDR_WD].height = 18
for rr in range(DATA_R0, tot_tech + 1):
    ws.row_dimensions[rr].height = 18

ws.sheet_properties.outlinePr.summaryRight = True
ws.sheet_properties.outlinePr.summaryBelow = False
ws.freeze_panes = "C4"
ws.sheet_view.zoomScale = 100

wb.active = wb.sheetnames.index("Численность по дням")

# Сохранение через ASCII temp (на всякий случай — путь содержит кириллицу)
with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
    tmp_path = tmp.name
wb.save(tmp_path)
os.makedirs(os.path.dirname(OUT_FINAL), exist_ok=True)
shutil.copyfile(tmp_path, OUT_FINAL)
os.remove(tmp_path)
print("OK:", OUT_FINAL)
print(f"  last_col={last_col} ({get_column_letter(last_col)})")
print(f"  rows: data {DATA_R0}..{tot_sum}")
