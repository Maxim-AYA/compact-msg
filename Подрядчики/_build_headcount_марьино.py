"""Сборка xlsx «Численность подрядчиков на площадке - Марьино» (ежедневный табель).

Структура идентична файлам Репино / Бугры-3 (см. `_build_headcount.py`,
`_build_headcount_бугры3.py`):
- Лист «Справочник» — 13 подрядчиков × организация + зона ответственности.
- Лист «Численность по дням» — 12 месяцев с мая 2026 по дням,
  после каждого месяца колонка «Средн. числ.», группировка по столбцам.
- Строки: 13 подрядчиков × (ИТР / Раб / Σ / Техн.) + ИТОГО × 4.

Список подрядчиков — из файла `Марьино_Документы подрядчиков.xlsx`
(группа «СК Марьино», file_id 1638006 на Битрикс-Диске).
Зоны ответственности — из строки 31 того же файла («Способы выполнения работ»).

Стартовые данные за 18-20.05.2026 — извлечены из Битрикс-чата группы
«Подрядчики Марьино» (chat_id 74200), куда подрядчики ежедневно пишут
утренние отчёты с численностью (ИТР / Раб / Техника). Чат активен
с 16.05.2026; до этого ежедневная отчётность шла в Telegram «Подрядчики
Марьино» (по Памятке подрядчика).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from datetime import date
import calendar, shutil, os, tempfile

OUT_FINAL = r"C:\Авраменко\Claude Code Projects\МСГ\Подрядчики\Марьино\Численность подрядчиков на площадке - Марьино.xlsx"

# === Подрядчики (по Марьино_Документы подрядчиков.xlsx, ред. 15.05.2026) ===
contractors = [
    ("АО СК «Компакт»",         "Генподрядчик"),
    ("ООО «СФЕРА»",             "Фасады (сэндвич-панели)"),
    ("ООО «ИСК ЭНЕРГОСИСТЕМА»", "Газопровод, земляные работы"),
    ("ООО «АВТОДОРКОМЛЕКС»",    "Благоустройство"),
    ("ООО «НВ-Строй»",          "Сети водоснабжения наружные"),
    ("ООО «ССК-26»",            "Монтаж металлоконструкций"),
    ("ООО «Гарант ПБ»",         "Слаботочка, пожарная сигнализация"),
    ("ООО «АСКОН»",             "Фасады, витражи"),
    ("ООО «ТВН»",               "Наружное электроснабжение"),
    ("ООО «СМК»",               "Отопление и вентиляция"),
    ("ООО «Электрокомплекс»",   "Слаботочка (внутри)"),
    ("ИП «Карпенко»",           "Каменная кладка"),
]
contr_short = [c[0] for c in contractors]
contr_zone  = [c[1] for c in contractors]
N_CONTR = len(contractors)

# === Стили ===
NAVY = "1F4E79"
LIGHT_BLUE = "DCE6F1"
LIGHTER_BLUE = "EAF1F8"
GREY_LIGHT = "F2F2F2"
WEEKEND_FILL = "FFF2CC"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
DATA_BOLD = Font(name="Calibri", size=10, bold=True)
THIN = Side(border_style="thin", color="A6A6A6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NW = Alignment(horizontal="center", vertical="center")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
SUM_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
SUB_FILL = PatternFill("solid", fgColor=LIGHTER_BLUE)
GREY_FILL = PatternFill("solid", fgColor=GREY_LIGHT)
WEEKEND = PatternFill("solid", fgColor=WEEKEND_FILL)
TECH_FILL = PatternFill("solid", fgColor="E2EFDA")

wb = Workbook()

# ============================================================
# Лист «Справочник»
# ============================================================
ws_ref = wb.active
ws_ref.title = "Справочник"
ws_ref.cell(row=1, column=1, value="№")
ws_ref.cell(row=1, column=2, value="Организация")
ws_ref.cell(row=1, column=3, value="Зона ответственности (СМР)")
for col in range(1, 4):
    c = ws_ref.cell(row=1, column=col)
    c.fill = NAVY_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

for i, (org, smr) in enumerate(contractors):
    r = i + 2
    ws_ref.cell(row=r, column=1, value=i + 1).alignment = CENTER_NW
    ws_ref.cell(row=r, column=2, value=org)
    ws_ref.cell(row=r, column=3, value=smr)
    for col in range(1, 4):
        c = ws_ref.cell(row=r, column=col)
        c.font = DATA_FONT
        c.border = BORDER
        if col > 1:
            c.alignment = Alignment(vertical="center")

ws_ref.column_dimensions["A"].width = 5
ws_ref.column_dimensions["B"].width = 30
ws_ref.column_dimensions["C"].width = 36
ws_ref.row_dimensions[1].height = 28
for r in range(2, 2 + N_CONTR):
    ws_ref.row_dimensions[r].height = 20

# ============================================================
# Лист «Численность по дням»
# ============================================================
ws = wb.create_sheet("Численность по дням")

# Период: 12 месяцев с мая 2026
months = []
y, m = 2026, 5
for _ in range(12):
    months.append((y, m))
    m += 1
    if m > 12:
        m = 1; y += 1

months_ru = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг",
             "Сен", "Окт", "Ноя", "Дек"]
wd_ru = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]

HDR_MONTH = 1
HDR_DAY = 2
HDR_WD = 3
DATA_R0 = 4

ws.cell(row=HDR_MONTH, column=1, value="Подрядчик")
ws.merge_cells(start_row=HDR_MONTH, start_column=1, end_row=HDR_WD, end_column=1)
ws.cell(row=HDR_MONTH, column=2, value="Кат.")
ws.merge_cells(start_row=HDR_MONTH, start_column=2, end_row=HDR_WD, end_column=2)
for col in (1, 2):
    c = ws.cell(row=HDR_MONTH, column=col)
    c.fill = NAVY_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

month_cols = {}
col = 3
for (yy, mm) in months:
    ndays = calendar.monthrange(yy, mm)[1]
    days = []
    for d in range(1, ndays + 1):
        wd = date(yy, mm, d).weekday()
        days.append((col, d, wd))
        col += 1
    total_col = col
    col += 1
    month_cols[(yy, mm)] = {"days": days, "total_col": total_col, "ndays": ndays}

last_col = col - 1

for (yy, mm), info in month_cols.items():
    start_col = info["days"][0][0]
    end_col = info["total_col"]

    ws.cell(row=HDR_MONTH, column=start_col, value=f"{months_ru[mm-1]} {yy}")
    ws.merge_cells(start_row=HDR_MONTH, start_column=start_col,
                   end_row=HDR_MONTH, end_column=end_col)
    mc = ws.cell(row=HDR_MONTH, column=start_col)
    mc.font = HEADER_FONT
    mc.alignment = CENTER
    mc.fill = NAVY_FILL
    mc.border = BORDER

    for c, d, wd in info["days"]:
        cell = ws.cell(row=HDR_DAY, column=c, value=d)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_NW
        cell.fill = NAVY_FILL
        cell.border = BORDER
    cell = ws.cell(row=HDR_DAY, column=info["total_col"], value="Средн.")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_NW
    cell.fill = NAVY_FILL
    cell.border = BORDER

    for c, d, wd in info["days"]:
        cell = ws.cell(row=HDR_WD, column=c, value=wd_ru[wd])
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_NW
        cell.fill = NAVY_FILL
        cell.border = BORDER
    cell = ws.cell(row=HDR_WD, column=info["total_col"], value="числ.")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_NW
    cell.fill = NAVY_FILL
    cell.border = BORDER

# Строки подрядчиков (N × 4)
contr_rows = {}
r = DATA_R0
bold_inline = InlineFont(rFont="Calibri", sz=10, b=True, color="000000")
reg_inline  = InlineFont(rFont="Calibri", sz=10, b=False, color="000000")
for i in range(N_CONTR):
    name = contr_short[i]
    zone = contr_zone[i]
    rich = CellRichText([
        TextBlock(bold_inline, name),
        TextBlock(reg_inline, f"\n({zone})"),
    ])
    ws.cell(row=r, column=1, value=rich)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=1)
    nc = ws.cell(row=r, column=1)
    nc.alignment = Alignment(vertical="center", horizontal="left",
                             wrap_text=True, indent=1)
    nc.fill = GREY_FILL
    nc.border = BORDER

    ws.cell(row=r,     column=2, value="ИТР")
    ws.cell(row=r + 1, column=2, value="Раб")
    ws.cell(row=r + 2, column=2, value="Σ")
    ws.cell(row=r + 3, column=2, value="Техн.")
    for offset, fill in ((0, SUB_FILL), (1, SUB_FILL), (2, SUM_FILL), (3, TECH_FILL)):
        cc = ws.cell(row=r + offset, column=2)
        cc.font = DATA_BOLD if offset >= 2 else DATA_FONT
        cc.alignment = CENTER_NW
        cc.border = BORDER
        cc.fill = fill

    contr_rows[name] = {"itr": r, "rab": r + 1, "sum": r + 2, "tech": r + 3}
    r += 4

# ИТОГО снизу
tot_itr  = r
tot_rab  = r + 1
tot_sum  = r + 2
tot_tech = r + 3
ws.cell(row=tot_itr, column=1, value="ИТОГО")
ws.merge_cells(start_row=tot_itr, start_column=1, end_row=tot_tech, end_column=1)
tc = ws.cell(row=tot_itr, column=1)
tc.font = HEADER_FONT
tc.alignment = CENTER
tc.fill = NAVY_FILL
tc.border = BORDER

ws.cell(row=tot_itr,  column=2, value="ИТР")
ws.cell(row=tot_rab,  column=2, value="Раб")
ws.cell(row=tot_sum,  column=2, value="Σ")
ws.cell(row=tot_tech, column=2, value="Техн.")
for rr, fill in ((tot_itr, SUM_FILL), (tot_rab, SUM_FILL),
                 (tot_sum, SUM_FILL), (tot_tech, TECH_FILL)):
    cc = ws.cell(row=rr, column=2)
    cc.font = DATA_BOLD
    cc.alignment = CENTER_NW
    cc.fill = fill
    cc.border = BORDER

# === Стартовые данные из Битрикс-чата 74200 «Подрядчики Марьино» ===
# tuple = (ИТР, Раб, Техн). Если значения нет в чате — подрядчик опускается.
chat_data = {
    (2026, 5, 18): {
        # Компакт: 2 ИТР, газобетонщики 5 + отделочники 2 + бригада Филимонова 8 = 15 раб
        "АО СК «Компакт»":         (2, 15, 0),
        "ООО «ССК-26»":            (1, 4,  1),
        "ООО «АСКОН»":             (3, 5,  0),
        "ООО «Электрокомплекс»":   (1, 6,  3),
        "ООО «Гарант ПБ»":         (1, 5,  4),
        "ООО «АВТОДОРКОМЛЕКС»":    (2, 6,  4),
        "ООО «НВ-Строй»":          (4, 16, 7),
        "ООО «ТВН»":               (1, 5,  2),
        "ООО «СМК»":               (1, 11, 0),
    },
    (2026, 5, 19): {
        "ООО «ССК-26»":            (1, 4,  1),
        "ООО «АСКОН»":             (3, 5,  0),
        "ООО «Электрокомплекс»":   (1, 6,  3),
        "ООО «АВТОДОРКОМЛЕКС»":    (0, 1,  1),
        "ООО «Гарант ПБ»":         (1, 5,  4),
        "ООО «НВ-Строй»":          (4, 16, 6),
    },
    (2026, 5, 20): {
        "ООО «АВТОДОРКОМЛЕКС»":    (0, 1,  1),
        "ООО «Электрокомплекс»":   (1, 2,  3),
        "ООО «ССК-26»":            (1, 4,  1),
        "ООО «АСКОН»":             (3, 5,  0),
        "ООО «Гарант ПБ»":         (1, 5,  4),
        "ООО «НВ-Строй»":          (3, 15, 7),
    },
}

# === Формулы Σ и средние + значения из чата ===
for name in contr_short:
    rows = contr_rows[name]
    itr_r, rab_r, sum_r, tech_r = rows["itr"], rows["rab"], rows["sum"], rows["tech"]

    for (yy, mm), info in month_cols.items():
        for c, d, wd in info["days"]:
            L = get_column_letter(c)
            day_key = (yy, mm, d)
            if day_key in chat_data and name in chat_data[day_key]:
                itr, rab, tech = chat_data[day_key][name]
                if itr: ws.cell(row=itr_r,  column=c, value=itr)
                if rab: ws.cell(row=rab_r,  column=c, value=rab)
                if tech: ws.cell(row=tech_r, column=c, value=tech)
            ws.cell(row=sum_r, column=c,
                    value=f'=IF(AND({L}{itr_r}="",{L}{rab_r}=""),"",N({L}{itr_r})+N({L}{rab_r}))')

        tc_col = info["total_col"]
        first_L = get_column_letter(info["days"][0][0])
        last_L  = get_column_letter(info["days"][-1][0])
        def avg_formula(row):
            return f'=IFERROR(ROUND(AVERAGEIF({first_L}{row}:{last_L}{row},">0"),1),"")'
        ws.cell(row=itr_r,  column=tc_col, value=avg_formula(itr_r))
        ws.cell(row=rab_r,  column=tc_col, value=avg_formula(rab_r))
        ws.cell(row=sum_r,  column=tc_col, value=avg_formula(sum_r))
        ws.cell(row=tech_r, column=tc_col, value=avg_formula(tech_r))

# ИТОГО строки
all_itr_rows  = [contr_rows[n]["itr"]  for n in contr_short]
all_rab_rows  = [contr_rows[n]["rab"]  for n in contr_short]
all_tech_rows = [contr_rows[n]["tech"] for n in contr_short]

for (yy, mm), info in month_cols.items():
    for c, _, _ in info["days"]:
        L = get_column_letter(c)
        itr_cells  = ",".join(f"{L}{rr}" for rr in all_itr_rows)
        rab_cells  = ",".join(f"{L}{rr}" for rr in all_rab_rows)
        tech_cells = ",".join(f"{L}{rr}" for rr in all_tech_rows)
        ws.cell(row=tot_itr,  column=c, value=f"=SUM({itr_cells})")
        ws.cell(row=tot_rab,  column=c, value=f"=SUM({rab_cells})")
        ws.cell(row=tot_sum,  column=c, value=f'=IF(AND({L}{tot_itr}="",{L}{tot_rab}=""),"",N({L}{tot_itr})+N({L}{tot_rab}))')
        ws.cell(row=tot_tech, column=c, value=f"=SUM({tech_cells})")
    tc_col = info["total_col"]
    first_L = get_column_letter(info["days"][0][0])
    last_L  = get_column_letter(info["days"][-1][0])
    for rr in (tot_itr, tot_rab, tot_sum, tot_tech):
        ws.cell(row=rr, column=tc_col,
                value=f'=IFERROR(ROUND(AVERAGEIF({first_L}{rr}:{last_L}{rr},">0"),1),"")')

# Стили ячеек данных + outline
sum_rows_set = {contr_rows[n]["sum"]  for n in contr_short} | {tot_sum}
tech_contr_rows = {contr_rows[n]["tech"] for n in contr_short}

for (yy, mm), info in month_cols.items():
    for c, d, wd in info["days"]:
        L = get_column_letter(c)
        ws.column_dimensions[L].outline_level = 1
        ws.column_dimensions[L].width = 4.2
        is_weekend = wd >= 5
        for rr in range(DATA_R0, tot_tech + 1):
            cell = ws.cell(row=rr, column=c)
            cell.font = DATA_BOLD if rr in sum_rows_set or rr >= tot_itr else DATA_FONT
            cell.alignment = CENTER_NW
            cell.border = BORDER
            if is_weekend:
                cell.fill = WEEKEND
            elif rr == tot_tech:
                cell.fill = TECH_FILL
            elif rr in tech_contr_rows:
                pass
            elif rr in sum_rows_set:
                cell.fill = SUB_FILL
            elif rr == tot_itr or rr == tot_rab:
                cell.fill = SUB_FILL
        if is_weekend:
            wdc = ws.cell(row=HDR_WD, column=c)
            wdc.fill = PatternFill("solid", fgColor="B85450")

    tc_col = info["total_col"]
    ws.column_dimensions[get_column_letter(tc_col)].width = 8
    for rr in range(DATA_R0, tot_tech + 1):
        cell = ws.cell(row=rr, column=tc_col)
        cell.fill = TECH_FILL if (rr in tech_contr_rows or rr == tot_tech) else SUM_FILL
        cell.font = DATA_BOLD
        cell.alignment = CENTER_NW
        cell.border = BORDER

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 5

for (yy, mm), info in month_cols.items():
    if yy == 2026 and mm == 5:
        continue
    for c, d, wd in info["days"]:
        ws.column_dimensions[get_column_letter(c)].hidden = True
    ws.column_dimensions[get_column_letter(info["total_col"])].collapsed = True

ws.row_dimensions[HDR_MONTH].height = 22
ws.row_dimensions[HDR_DAY].height = 18
ws.row_dimensions[HDR_WD].height = 18
for rr in range(DATA_R0, tot_tech + 1):
    ws.row_dimensions[rr].height = 18

ws.sheet_properties.outlinePr.summaryRight = True
ws.sheet_properties.outlinePr.summaryBelow = False
ws.freeze_panes = "C4"
ws.sheet_view.zoomScale = 100

wb.active = wb.sheetnames.index("Численность по дням")

with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
    tmp_path = tmp.name
wb.save(tmp_path)
os.makedirs(os.path.dirname(OUT_FINAL), exist_ok=True)
shutil.copyfile(tmp_path, OUT_FINAL)
os.remove(tmp_path)
print("OK:", OUT_FINAL)
print(f"  contractors: {N_CONTR}")
print(f"  last_col={last_col} ({get_column_letter(last_col)})")
print(f"  rows: data {DATA_R0}..{tot_sum} (tot_tech={tot_tech})")
